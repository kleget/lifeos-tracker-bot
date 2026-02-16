from __future__ import annotations

import asyncio
import json
import logging
import sys
import threading
from pathlib import Path
from datetime import datetime, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import Workbook

from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.error import BadRequest
from telegram.ext import (
    ApplicationBuilder,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

from config import load_config
from menus import (
    MAIN_MENU,
    SPORT_MENU,
    TRAINING_OPTIONS,
    CARDIO_OPTIONS,
    STEPS_OPTIONS,
    STUDY_MENU,
    ENGLISH_OPTIONS,
    ML_OPTIONS,
    ALGOS_OPTIONS,
    UNI_OPTIONS,
    CODE_MODE_OPTIONS,
    CODE_TOPIC_OPTIONS,
    READING_OPTIONS,
    LEISURE_MENU,
    REST_TIME_OPTIONS,
    REST_TYPE_OPTIONS,
    SLEEP_BEDTIME_OPTIONS,
    SLEEP_HOURS_OPTIONS,
    SLEEP_REGIME_OPTIONS,
    PRODUCTIVITY_OPTIONS,
    PROCRASTINATION_OPTIONS,
    EXPENSE_OPTIONS,
    FOOD_MENU,
    FOOD_PROTEIN_OPTIONS,
    FOOD_GARNISH_OPTIONS,
    FOOD_SWEET_OPTIONS,
    FOOD_OIL_OPTIONS,
    MORALE_MENU,
    MOOD_OPTIONS,
    ENERGY_OPTIONS,
    HABITS_MENU,
    build_keyboard,
    quantity_keyboard,
    NAP_OPTIONS,
)
from db import Database

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s %(message)s")
LOGGER = logging.getLogger("lifeos-bot")

BASE_DIR = Path(__file__).resolve().parent.parent

DAILY_HEADERS = [
    "Дата",
    "Тренировка",
    "Кардио_мин",
    "Шаги_категория",
    "Шаги_кол-во",
    "Английский_мин",
    "ML_мин",
    "Алгосы_мин",
    "ВУЗ_мин",
    "Код_режим",
    "Код_тема",
    "Чтение_стр",
    "Отдых_время",
    "Отдых_тип",
    "Сон_отбой",
    "Сон_часы",
    "Сон_дневной",
    "Режим",
    "Продуктивность",
    "Стрельнул_раз",
    "Настроение",
    "Энергия",
    "Траты_всего",
    "Траты_еда",
    "Траты_одежда",
    "Траты_бытовуха",
    "Траты_гульки",
    "Траты_здоровье",
    "Траты_другое",
    "Вес",
    "О_чем_жалею",
    "Отзыв_о_дне",
    "Привычки",
    "Активные_ккал",
    "Еда_учтена",
    "Еда_ккал",
    "Еда_Б",
    "Еда_Ж",
    "Еда_У",
    "Еда_источник",
    "Качество_дня",
    "Коэффициент_дня",
    "Не_заполнено",
]

COLUMN_MAP = {
    "training": "training",
    "cardio": "cardio_min",
    "steps": "steps_category",
    "steps_count": "steps_count",
    "english": "english_min",
    "ml": "ml_min",
    "algos": "algo_min",
    "uni": "uni_min",
    "code_mode": "code_mode",
    "code_topic": "code_topic",
    "reading": "reading_pages",
    "rest_time": "rest_time",
    "rest_type": "rest_type",
    "sleep_bed": "sleep_bed",
    "sleep_hours": "sleep_hours",
    "nap": "nap_hours",
    "sleep_regime": "sleep_regime",
    "productivity": "productivity",
    "shots": "shots_count",
    "mood": "mood",
    "energy": "energy",
    "weight": "weight",
    "regret": "regret",
    "review": "review",
    "habits": "habits",
    "active_kcal": "active_kcal",
    "food_tracked": "food_tracked",
    "food_kcal": "food_kcal",
    "food_protein": "food_protein",
    "food_fat": "food_fat",
    "food_carb": "food_carb",
    "food_source": "food_source",
}

DB_TO_HEADER = {
    "training": "Тренировка",
    "cardio_min": "Кардио_мин",
    "steps_category": "Шаги_категория",
    "steps_count": "Шаги_кол-во",
    "english_min": "Английский_мин",
    "ml_min": "ML_мин",
    "algo_min": "Алгосы_мин",
    "uni_min": "ВУЗ_мин",
    "code_mode": "Код_режим",
    "code_topic": "Код_тема",
    "reading_pages": "Чтение_стр",
    "rest_time": "Отдых_время",
    "rest_type": "Отдых_тип",
    "sleep_bed": "Сон_отбой",
    "sleep_hours": "Сон_часы",
    "nap_hours": "Сон_дневной",
    "sleep_regime": "Режим",
    "productivity": "Продуктивность",
    "shots_count": "Стрельнул_раз",
    "mood": "Настроение",
    "energy": "Энергия",
    "weight": "Вес",
    "regret": "О_чем_жалею",
    "review": "Отзыв_о_дне",
    "habits": "Привычки",
    "active_kcal": "Активные_ккал",
    "food_tracked": "Еда_учтена",
    "food_kcal": "Еда_ккал",
    "food_protein": "Еда_Б",
    "food_fat": "Еда_Ж",
    "food_carb": "Еда_У",
    "food_source": "Еда_источник",
}

NUMERIC_FIELDS = {"cardio", "english", "ml", "algos", "uni", "reading", "productivity"}

EXPENSE_CATEGORY_LABELS = {
    "food": "Еда",
    "clothes": "Одежда",
    "household": "Бытовуха",
    "party": "Гульки",
    "health": "Здоровье",
    "other": "Другое",
}

EXPENSE_HEADER_BY_LABEL = {
    "Еда": "Траты_еда",
    "Одежда": "Траты_одежда",
    "Бытовуха": "Траты_бытовуха",
    "Гульки": "Траты_гульки",
    "Здоровье": "Траты_здоровье",
    "Другое": "Траты_другое",
}


def get_now(tz_name: str) -> datetime:
    try:
        return datetime.now(ZoneInfo(tz_name))
    except ZoneInfoNotFoundError:
        LOGGER.warning("Timezone '%s' not found. Falling back to local time.", tz_name)
        return datetime.now()


def today_str(tz_name: str) -> str:
    return get_now(tz_name).strftime("%Y-%m-%d")


def time_str(tz_name: str) -> str:
    return get_now(tz_name).strftime("%H:%M")


def get_sheets(context: ContextTypes.DEFAULT_TYPE) -> Database:
    return context.application.bot_data["db"]


STATE_ACTIVE_DAY = "active_day"
STATE_SLEEP_START = "sleep_start"
STATE_SLEEP_START_DAY = "sleep_start_day"
STATE_SLEEP_START_BED = "sleep_start_bed"
STATE_VIEW_DATE = "view_date"


def get_active_date(context: ContextTypes.DEFAULT_TYPE) -> str:
    cfg = context.application.bot_data["config"]
    db = get_sheets(context)
    today = today_str(cfg.timezone)
    active = db.get_state(STATE_ACTIVE_DAY)
    if not active:
        db.set_state(STATE_ACTIVE_DAY, today)
        return today
    return active


def get_view_date(context: ContextTypes.DEFAULT_TYPE) -> str:
    selected = context.user_data.get(STATE_VIEW_DATE)
    if selected:
        return str(selected)
    return get_active_date(context)


def set_view_date(context: ContextTypes.DEFAULT_TYPE, date_str: str | None) -> None:
    if date_str:
        context.user_data[STATE_VIEW_DATE] = date_str
    else:
        context.user_data.pop(STATE_VIEW_DATE, None)


def get_sleep_start(context: ContextTypes.DEFAULT_TYPE) -> datetime | None:
    db = get_sheets(context)
    raw = db.get_state(STATE_SLEEP_START)
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw)
    except ValueError:
        return None


def summary_state_key(chat_id: int) -> str:
    return f"summary_msg_{chat_id}"


def prompt_state_key(chat_id: int) -> str:
    return f"prompt_msg_{chat_id}"


def get_state_int(db: Database, key: str) -> int | None:
    raw = db.get_state(key)
    if not raw:
        return None
    try:
        return int(raw)
    except ValueError:
        return None


async def safe_delete_message(bot, chat_id: int, message_id: int | None) -> None:
    if not message_id:
        return
    try:
        await bot.delete_message(chat_id, message_id)
    except Exception:
        return


async def send_or_edit_summary(
    context: ContextTypes.DEFAULT_TYPE,
    chat_id: int,
    text: str,
    keyboard: InlineKeyboardMarkup,
) -> None:
    db = get_sheets(context)
    msg_id = get_state_int(db, summary_state_key(chat_id))
    if msg_id:
        try:
            await context.bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=keyboard)
            return
        except BadRequest as exc:
            if "message is not modified" in str(exc).lower():
                return
        except Exception:
            pass
        await safe_delete_message(context.bot, chat_id, msg_id)
    sent = await context.bot.send_message(chat_id=chat_id, text=text, reply_markup=keyboard)
    db.set_state(summary_state_key(chat_id), str(sent.message_id))


async def send_or_edit_prompt(
    context: ContextTypes.DEFAULT_TYPE,
    chat_id: int,
    text: str,
    keyboard: InlineKeyboardMarkup | None = None,
) -> int:
    db = get_sheets(context)
    msg_id = get_state_int(db, prompt_state_key(chat_id))
    if msg_id:
        try:
            await context.bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=keyboard)
            return msg_id
        except Exception:
            pass
    sent = await context.bot.send_message(chat_id=chat_id, text=text, reply_markup=keyboard)
    db.set_state(prompt_state_key(chat_id), str(sent.message_id))
    return sent.message_id


async def clear_prompt(context: ContextTypes.DEFAULT_TYPE, chat_id: int) -> None:
    db = get_sheets(context)
    msg_id = get_state_int(db, prompt_state_key(chat_id))
    if msg_id:
        await safe_delete_message(context.bot, chat_id, msg_id)
        db.set_state(prompt_state_key(chat_id), None)


async def render_summary(context: ContextTypes.DEFAULT_TYPE, chat_id: int, date_str: str | None = None) -> None:
    if date_str is None:
        date_str = get_active_date(context)
    summary = await build_daily_summary(context, date_str)
    daily = get_daily_data(context, date_str)
    await send_or_edit_summary(context, chat_id, summary, build_main_menu_keyboard(daily))


async def safe_render_summary(context: ContextTypes.DEFAULT_TYPE, chat_id: int, date_str: str | None = None) -> None:
    try:
        await render_summary(context, chat_id, date_str)
    except Exception:
        LOGGER.exception("Failed to render summary")
        db = get_sheets(context)
        set_view_date(context, None)
        fallback_date = get_active_date(context)
        db.ensure_daily_row(fallback_date)
        daily = get_daily_data(context, fallback_date)
        text = f"📅 Сегодня: {fallback_date}\nПока нет данных."
        await send_or_edit_summary(context, chat_id, text, build_main_menu_keyboard(daily))


async def finalize_input(context: ContextTypes.DEFAULT_TYPE, chat_id: int, user_message_id: int) -> None:
    await clear_prompt(context, chat_id)
    await safe_delete_message(context.bot, chat_id, user_message_id)
    await safe_render_summary(context, chat_id, get_view_date(context))


def build_stats_keyboard(selected: str) -> InlineKeyboardMarkup:
    rows = [
        [
            InlineKeyboardButton(("✅ " if selected == "all" else "") + "Все", callback_data="stats:all"),
            InlineKeyboardButton(("✅ " if selected == "month" else "") + "30д", callback_data="stats:month"),
            InlineKeyboardButton(("✅ " if selected == "week" else "") + "7д", callback_data="stats:week"),
        ],
        [InlineKeyboardButton("⬅️ К сводке", callback_data="stats:back")],
    ]
    return InlineKeyboardMarkup(rows)


def avg_value(total: float, count: int, digits: int = 0) -> str:
    if count <= 0:
        return "—"
    return fmt_num(total / count, digits)


def stats_period_dates(dates: list[str], period: str, tz_name: str) -> tuple[str, list[str]]:
    today = get_now(tz_name).date()
    if period == "week":
        start = today - timedelta(days=6)
        label = "7 дней"
    elif period == "month":
        start = today - timedelta(days=29)
        label = "30 дней"
    else:
        start = None
        label = "всё время"
    if start is None:
        return label, dates
    start_str = start.isoformat()
    return label, [d for d in dates if d >= start_str]


def shooting_activity_label(last7_total: int, last7_shot_days: int) -> str:
    if last7_shot_days == 7 and last7_total >= 21:
        return "супер активно"
    if last7_total >= 12 and last7_shot_days >= 4:
        return "активно"
    if last7_total >= 6 and last7_shot_days >= 3:
        return "умеренно"
    if last7_total >= 2:
        return "очень мало"
    if last7_total == 1:
        return "единично"
    return "нет активности"


def rolling_window_max(values: list[int], window: int) -> int:
    if not values:
        return 0
    if window <= 1:
        return max(values)
    current = sum(values[:window])
    best = current
    for idx in range(window, len(values)):
        current += values[idx] - values[idx - window]
        if current > best:
            best = current
    return best


def build_stats_summary(context: ContextTypes.DEFAULT_TYPE, period: str) -> str:
    cfg = context.application.bot_data["config"]
    db = get_sheets(context)
    all_dates = db.get_daily_dates()
    label, dates = stats_period_dates(all_dates, period, cfg.timezone)
    if not dates:
        return f"📊 Статистика ({label})\nНет данных."

    full = partial = none = tracked = 0
    quality_sum = quality_count = 0
    sleep_sum = sleep_count = 0
    steps_sum = steps_count = 0
    kcal_sum = prot_sum = fat_sum = carb_sum = 0.0
    kbju_count = 0
    eng_sum = eng_count = 0
    ml_sum = ml_count = 0
    alg_sum = alg_count = 0
    uni_sum = uni_count = 0
    shots_period_total = shots_period_days = 0
    shots_by_date: dict[str, int] = {}
    last_shot_date: str | None = None
    expense_total_sum = 0.0
    expense_days = 0
    expense_max_day = 0.0
    expense_max_date = ""
    expense_by_category: dict[str, float] = {label: 0.0 for label in EXPENSE_HEADER_BY_LABEL}

    for date_str in all_dates:
        data = get_daily_data(context, date_str)
        shots = int(parse_sheet_number(data.get("Стрельнул_раз")))
        shots_by_date[date_str] = shots
        if shots > 0:
            last_shot_date = date_str
        if date_str in dates and shots > 0:
            shots_period_days += 1
        if date_str in dates:
            shots_period_total += shots

    for date_str in dates:
        data = get_daily_data(context, date_str)
        status = day_completion_status(data)
        if status == "empty":
            continue
        tracked += 1
        if status == "full":
            full += 1
        elif status == "partial":
            partial += 1
        else:
            none += 1

        quality = compute_quality(data)
        if quality is not None:
            quality_sum += quality
            quality_count += 1

        sleep = parse_sleep_hours(data.get("Сон_часы")) or 0.0
        nap = parse_sheet_number(data.get("Сон_дневной"))
        sleep_total = sleep + max(0.0, nap)
        if sleep_total > 0:
            sleep_sum += sleep_total
            sleep_count += 1

        steps = steps_value(data)
        if steps > 0:
            steps_sum += steps
            steps_count += 1

        kcal = data.get("Еда_ккал")
        protein = data.get("Еда_Б")
        fat = data.get("Еда_Ж")
        carbs = data.get("Еда_У")
        if any(is_set(v) for v in (kcal, protein, fat, carbs)):
            kcal_sum += parse_sheet_number(kcal)
            prot_sum += parse_sheet_number(protein)
            fat_sum += parse_sheet_number(fat)
            carb_sum += parse_sheet_number(carbs)
            kbju_count += 1

        english = int(parse_sheet_number(data.get("Английский_мин")))
        if english > 0:
            eng_sum += english
            eng_count += 1
        ml = int(parse_sheet_number(data.get("ML_мин")))
        if ml > 0:
            ml_sum += ml
            ml_count += 1
        algos = int(parse_sheet_number(data.get("Алгосы_мин")))
        if algos > 0:
            alg_sum += algos
            alg_count += 1
        uni = int(parse_sheet_number(data.get("ВУЗ_мин")))
        if uni > 0:
            uni_sum += uni
            uni_count += 1

        day_expense = parse_sheet_number(data.get("Траты_всего"))
        if day_expense > 0:
            expense_days += 1
            if day_expense > expense_max_day:
                expense_max_day = day_expense
                expense_max_date = date_str
        expense_total_sum += day_expense
        for label, header in EXPENSE_HEADER_BY_LABEL.items():
            expense_by_category[label] += parse_sheet_number(data.get(header))

    if tracked == 0:
        return f"📊 Статистика ({label})\nНет данных."

    full_pct = int(round(full / tracked * 100))
    partial_pct = int(round(partial / tracked * 100))
    none_pct = int(round(none / tracked * 100))

    avg_quality = avg_value(quality_sum, quality_count, 0)
    avg_sleep = avg_value(sleep_sum, sleep_count, 1)
    avg_steps = avg_value(steps_sum, steps_count, 0)
    if kbju_count:
        kbju_line = (
            f"{avg_value(kcal_sum, kbju_count, 0)} | Б {avg_value(prot_sum, kbju_count, 0)} | "
            f"Ж {avg_value(fat_sum, kbju_count, 0)} | У {avg_value(carb_sum, kbju_count, 0)}"
        )
    else:
        kbju_line = "—"

    study_line = (
        f"англ {avg_value(eng_sum, eng_count, 0)}м · "
        f"ml {avg_value(ml_sum, ml_count, 0)}м · "
        f"алг {avg_value(alg_sum, alg_count, 0)}м · "
        f"вуз {avg_value(uni_sum, uni_count, 0)}м"
    )

    period_span = len(dates)
    shot_freq_pct = int(round((shots_period_days / period_span) * 100)) if period_span > 0 else 0
    shots_avg_day = fmt_num(shots_period_total / period_span, 1) if period_span > 0 else "—"

    active_day = db.get_state(STATE_ACTIVE_DAY) or today_str(cfg.timezone)
    if last_shot_date:
        days_no_shot = max(
            0,
            (datetime.fromisoformat(active_day).date() - datetime.fromisoformat(last_shot_date).date()).days,
        )
    else:
        days_no_shot = period_span

    if all_dates:
        series_start = datetime.fromisoformat(all_dates[0]).date()
        series_end = datetime.fromisoformat(active_day).date()
        span = (series_end - series_start).days + 1
        by_day = [0] * max(1, span)
        for d, value in shots_by_date.items():
            idx = (datetime.fromisoformat(d).date() - series_start).days
            if 0 <= idx < len(by_day):
                by_day[idx] = value
        max_week = rolling_window_max(by_day, 7)
        max_month = rolling_window_max(by_day, 30)
        max_all = sum(by_day)
    else:
        max_week = max_month = max_all = 0

    last7_start = datetime.fromisoformat(active_day).date() - timedelta(days=6)
    last7_total = 0
    last7_days = 0
    for i in range(7):
        day = (last7_start + timedelta(days=i)).isoformat()
        value = shots_by_date.get(day, 0)
        last7_total += value
        if value > 0:
            last7_days += 1
    shoot_activity = shooting_activity_label(last7_total, last7_days)

    expense_avg_day = expense_total_sum / period_span if period_span > 0 else 0.0
    expense_avg_spend_day = expense_total_sum / expense_days if expense_days > 0 else 0.0
    expense_cat_parts: list[str] = []
    if expense_total_sum > 0:
        ranked = sorted(expense_by_category.items(), key=lambda x: x[1], reverse=True)
        for label_name, value in ranked:
            if value <= 0:
                continue
            share = int(round((value / expense_total_sum) * 100))
            expense_cat_parts.append(f"{label_name.lower()} {fmt_money(value)}₽ ({share}%)")
            if len(expense_cat_parts) >= 4:
                break

    lines = [
        f"📊 Статистика ({label})",
        f"Дней в выборке: {tracked}",
        f"✅ Полных: {full_pct}% ({full})",
        f"🟧 Частичных: {partial_pct}% ({partial})",
        f"❌ Провалов: {none_pct}% ({none})",
        f"⭐ Качество ср.: {avg_quality}",
        f"😴 Сон ср.: {avg_sleep} ч",
        f"🚶 Шаги ср.: {avg_steps}",
        f"🍽 КБЖУ ср.: {kbju_line}",
        f"📚 Учеба ср.: {study_line}",
        "",
        f"🎯 Стрельба: {shoot_activity}",
        f"• Частота: {shot_freq_pct}% дней ({shots_period_days}/{period_span})",
        f"• Ср. за день: {shots_avg_day} раз",
        f"• Без стрельбы: {days_no_shot} дн",
        f"• Пик: 7д {max_week} | 30д {max_month} | всё {max_all}",
        "",
        "💸 Траты",
        f"• Итого: {fmt_money(expense_total_sum)} ₽",
        f"• Ср. в день: {fmt_money(expense_avg_day)} ₽",
        f"• Дней с тратами: {expense_days}/{period_span}",
    ]
    if expense_days > 0:
        lines.append(f"• Ср. в день с тратами: {fmt_money(expense_avg_spend_day)} ₽")
    if expense_max_day > 0:
        lines.append(f"• Макс. за день: {fmt_money(expense_max_day)} ₽ ({expense_max_date})")
    if expense_cat_parts:
        lines.append(f"• По категориям: {' · '.join(expense_cat_parts)}")
    return "\n".join(lines)


async def render_stats(context: ContextTypes.DEFAULT_TYPE, chat_id: int, period: str = "week") -> None:
    text = build_stats_summary(context, period)
    await send_or_edit_summary(context, chat_id, text, build_stats_keyboard(period))


def build_main_menu_keyboard(data: dict) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    viewing_date = str(data.get("Дата") or "")
    active_date = str(data.get("_active_day") or "")
    if data.get("_sleep_start"):
        rows.append([InlineKeyboardButton("☀️ Я проснулся", callback_data="sleep:toggle")])
        rows.append(
            [
                InlineKeyboardButton("✏️ Исправить отбой", callback_data="sleep:edit"),
                InlineKeyboardButton("↩️ Отменить сон", callback_data="sleep:cancel"),
            ]
        )
        rows.append([InlineKeyboardButton("📅 Дата", callback_data="menu:date")])
        if active_date and viewing_date and viewing_date != active_date:
            rows.append([InlineKeyboardButton("↩️ К текущему дню", callback_data="date:today")])
        return InlineKeyboardMarkup(rows)

    rows.append([InlineKeyboardButton("😴 Лёг спать", callback_data="sleep:toggle")])
    rows.append(
        [
            InlineKeyboardButton("🔄 Обновить", callback_data="menu:refresh"),
            InlineKeyboardButton("📊 Статистика", callback_data="stats:week"),
        ]
    )
    rows.append([InlineKeyboardButton("📅 Дата", callback_data="menu:date")])
    if active_date and viewing_date and viewing_date != active_date:
        rows.append([InlineKeyboardButton("↩️ К текущему дню", callback_data="date:today")])
    row: list[InlineKeyboardButton] = []
    for label, payload in MAIN_MENU:
        row.append(InlineKeyboardButton(label, callback_data=payload))
        if len(row) >= 2:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    return InlineKeyboardMarkup(rows)


def build_shots_keyboard(count: int) -> InlineKeyboardMarkup:
    buttons = [
        ("➖", "shots:-"),
        ("➕", "shots:+"),
    ]
    return build_keyboard(buttons, cols=2, back=("⬅️ Назад", "menu:leisure"))


def is_authorized(context: ContextTypes.DEFAULT_TYPE, user_id: int | None) -> bool:
    allowed = context.application.bot_data.get("allowed_user_id")
    if not allowed:
        return True
    return user_id == allowed


async def handle_error(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    LOGGER.exception("Unhandled error: %s", context.error)
    if update is None:
        return
    try:
        if hasattr(update, "effective_chat") and update.effective_chat:
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="⚠️ Ошибка при обращении к базе. Попробуй ещё раз через минуту.",
            )
    except Exception:
        LOGGER.exception("Failed to send error message")


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_authorized(context, update.effective_user.id if update.effective_user else None):
        return
    cfg = context.application.bot_data["config"]
    set_view_date(context, None)
    date_str = get_active_date(context)
    db = get_sheets(context)
    db.ensure_daily_row(date_str)
    if update.message is None:
        return
    chat_id = update.effective_chat.id
    old_summary_id = get_state_int(db, summary_state_key(chat_id))
    if old_summary_id:
        await safe_delete_message(context.bot, chat_id, old_summary_id)
        db.set_state(summary_state_key(chat_id), None)
    await safe_render_summary(context, chat_id, date_str)
    await safe_delete_message(context.bot, update.effective_chat.id, update.message.message_id)


async def show_menu(query, title: str, buttons, back_to: str = "menu:main", cols: int = 2) -> None:
    await query.answer()
    await query.edit_message_text(title, reply_markup=build_keyboard(buttons, cols=cols, back=("⬅️ Назад", back_to)))


def parse_number(value: str) -> float:
    value = value.replace(",", ".").strip()
    return float(value)


def parse_time_hhmm(value: str) -> tuple[int, int]:
    text = value.strip()
    if ":" not in text:
        raise ValueError("time")
    hours_str, minutes_str = text.split(":", 1)
    if not hours_str.isdigit() or not minutes_str.isdigit():
        raise ValueError("time")
    hours = int(hours_str)
    minutes = int(minutes_str)
    if hours < 0 or hours > 23 or minutes < 0 or minutes > 59:
        raise ValueError("time")
    return hours, minutes


def parse_numbers(text: str, count: int) -> list[float]:
    parts = [p for p in text.replace(",", ".").split() if p.strip()]
    if len(parts) != count:
        raise ValueError
    return [float(p) for p in parts]


def steps_to_category(steps: float) -> str:
    if steps < 5000:
        return "<5k"
    if steps < 7000:
        return "5-7k"
    if steps < 10000:
        return "7-10k"
    if steps < 12000:
        return "10-12k"
    if steps < 15000:
        return "12-15k"
    return "15k+"


STEPS_CATEGORY_TO_COUNT = {
    "<5k": 4000,
    "5-7k": 6000,
    "7-10k": 8500,
    "10-12k": 11000,
    "12-15k": 13500,
    "15k+": 15000,
}


def steps_value(data: dict) -> float:
    steps_count = parse_sheet_number(data.get("Шаги_кол-во"))
    if steps_count > 0:
        return steps_count
    category = normalize_choice(data.get("Шаги_категория"))
    return float(STEPS_CATEGORY_TO_COUNT.get(category, 0))


def parse_sleep_hours(value: object) -> float | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if text in {"<6", "мало"}:
        return 5.5
    if text in {"6-8"}:
        return 7.0
    if text in {">8"}:
        return 8.5
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None

def parse_sheet_number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(" ", "").replace(",", ".")
    num = ""
    for ch in text:
        if ch.isdigit() or ch in {".", "-"}:
            num += ch
        elif num:
            break
    try:
        return float(num)
    except ValueError:
        return 0.0


def fmt_num(value: float, digits: int = 0) -> str:
    if digits <= 0:
        return str(int(round(value)))
    formatted = f"{value:.{digits}f}".rstrip("0").rstrip(".")
    return formatted.replace(".", ",")


def fmt_money(value: float) -> str:
    rounded = round(value, 2)
    if abs(rounded - round(rounded)) < 1e-6:
        return str(int(round(rounded)))
    return f"{rounded:.2f}".rstrip("0").rstrip(".").replace(".", ",")


def fmt_steps(value: float) -> str:
    if value <= 0:
        return "—"
    return f"{int(round(value)):,}".replace(",", " ")


def steps_status_square(steps: float) -> str:
    if steps <= 0:
        return "⬜"
    if steps < 6000:
        return "🟥"
    if steps < 8000:
        return "🟧"
    if steps < 10000:
        return "🟨"
    if steps < 12000:
        return "🟩"
    if steps < 15000:
        return "🟦"
    return "🟪"


def fmt_value(value: object) -> str:
    return str(value) if value not in (None, "") else "—"


def display_training(value: object) -> str | None:
    if value in (None, ""):
        return None
    if value == "Ноги":
        return "Низ"
    return str(value)


def reading_is_set(value: object) -> bool:
    return value not in (None, "")


def format_reading_label(value: object) -> str:
    if value in (None, ""):
        return "Чтение"
    text = normalize_choice(value)
    if text in {"0", "0.0"}:
        return "Чтение: не читал"
    return f"Чтение: {text} стр"


def sleep_toggle_label(data: dict) -> str:
    raw = data.get("_sleep_start")
    if not raw:
        return "😴 Лег спать"
    try:
        start_dt = datetime.fromisoformat(str(raw))
        time_label = start_dt.strftime("%H:%M")
        return f"☀️ Проснулся (c {time_label})"
    except ValueError:
        return "☀️ Проснулся"




def day_targets(training_value: str | None) -> dict | None:
    if training_value in {"Ноги", "Низ", "Верх"}:
        return {
            "label": "Тренировочный день",
            "kcal": (1900, 2000),
            "protein": (125, 135),
            "fat": (55, 65),
            "carb": (180, 210),
        }
    if training_value in {"Отдых", "Пропуск", "Пропустил"}:
        return {
            "label": "День отдыха",
            "kcal": (1700, 1800),
            "protein": (120, 130),
            "fat": (55, 65),
            "carb": (140, 170),
        }
    return None


TRAINING_SCORES = {
    "Пропуск": 0,
    "Пропустил": 0,
    "Отдых": 0.6,
    "Ноги": 1,
    "Верх": 1,
    "Низ": 1,
}

STEPS_SCORES = {
    "<5k": 0,
    "5-7k": 0.25,
    "7-10k": 0.5,
    "10-12k": 0.75,
    "12-15k": 0.9,
    "15k+": 1,
}

SLEEP_HOURS_SCORES = {"<6": 0, "6-8": 1, ">8": 0.8}
REGIME_SCORES = {"сбит": 0.3, "не сбит": 1}
MOOD_SCORES = {
    "Отличное": 1,
    "Веселый": 0.85,
    "Обычное": 0.7,
    "Серьезный": 0.6,
    "Раздраженный": 0.4,
    "Беспокойный": 0.3,
    "Злой": 0.2,
}
ENERGY_SCORES = {"нет": 0, "мало": 0.33, "есть": 0.66, "я живчик": 1}


def is_set(value: object) -> bool:
    return value not in (None, "")


def score_range(value: float, ranges: list[tuple[float, float, float]]) -> float:
    for min_val, max_val, score in ranges:
        if min_val <= value <= max_val:
            return score
    return 0.2


def score_kbju(training_value: str | None, macros: dict) -> list[float]:
    kcal = macros["kcal"]
    protein = macros["protein"]
    fat = macros["fat"]
    carb = macros["carb"]

    training_day = training_value in {"Ноги", "Верх", "Низ", "Фулл"}

    if training_day:
        kcal_ranges = [(1900, 2050, 1), (1850, 1899, 0.7), (2051, 2100, 0.7)]
        protein_ranges = [
            (125, 135, 1),
            (115, 124, 0.7),
            (136, 145, 0.7),
            (105, 114, 0.4),
            (146, 155, 0.4),
        ]
        carb_ranges = [
            (180, 210, 1),
            (160, 179, 0.7),
            (211, 230, 0.7),
            (140, 159, 0.4),
            (231, 250, 0.4),
        ]
    else:
        kcal_ranges = [(1700, 1850, 1), (1650, 1699, 0.7), (1851, 1900, 0.7)]
        protein_ranges = [
            (120, 130, 1),
            (110, 119, 0.7),
            (131, 140, 0.7),
            (100, 109, 0.4),
            (141, 150, 0.4),
        ]
        carb_ranges = [
            (140, 170, 1),
            (120, 139, 0.7),
            (171, 190, 0.7),
            (100, 119, 0.4),
            (191, 210, 0.4),
        ]

    fat_ranges = [
        (55, 65, 1),
        (50, 54, 0.7),
        (66, 70, 0.7),
        (45, 49, 0.4),
        (71, 75, 0.4),
    ]

    kcal_score = score_range(kcal, kcal_ranges)
    protein_score = score_range(protein, protein_ranges)
    fat_score = score_range(fat, fat_ranges)
    carb_score = score_range(carb, carb_ranges)
    return [kcal_score, protein_score, fat_score, carb_score]


def bonus_linear(value: float, min_val: float, max_val: float, max_bonus: float) -> float:
    if value <= min_val:
        return 0.0
    if value >= max_val:
        return max_bonus
    return (value - min_val) / (max_val - min_val) * max_bonus


def compute_quality(data: dict) -> int | None:
    min_ok, context = day_minimum_met(data)
    if context["any_data"] is False:
        return 0

    def segment(value: float, start: float, end: float, s0: float, s1: float) -> float:
        if value <= start:
            return s0
        if value >= end:
            return s1
        return s0 + (value - start) / (end - start) * (s1 - s0)

    sleep = context["sleep_hours"]
    if sleep <= 4:
        sleep_score = 0.0
    elif sleep <= 5:
        sleep_score = segment(sleep, 4.0, 5.0, 0.0, 0.2)
    elif sleep <= 6:
        sleep_score = segment(sleep, 5.0, 6.0, 0.2, 0.5)
    elif sleep <= 9:
        sleep_score = segment(sleep, 6.0, 9.0, 0.5, 1.0)
    else:
        sleep_score = 1.0

    steps = context["steps"]
    if steps <= 6000:
        steps_score = segment(steps, 0.0, 6000.0, 0.0, 0.5)
    elif steps <= 12000:
        steps_score = segment(steps, 6000.0, 12000.0, 0.5, 1.0)
    else:
        steps_score = 1.0

    english = context["english"]
    if english <= 30:
        english_score = segment(english, 0.0, 30.0, 0.0, 0.5)
    elif english <= 60:
        english_score = segment(english, 30.0, 60.0, 0.5, 1.0)
    else:
        english_score = 1.0

    deep = context["study_total"]
    if deep <= 60:
        deep_score = segment(deep, 0.0, 60.0, 0.0, 0.5)
    elif deep <= 120:
        deep_score = segment(deep, 60.0, 120.0, 0.5, 1.0)
    else:
        deep_score = 1.0

    uni = context["uni"]
    uni_bonus = bonus_linear(uni, 30.0, 180.0, 15.0)

    training = context["training"]
    if training in {"Верх", "Ноги", "Низ", "Фулл"}:
        sport_score = 1.0
    elif training == "Отдых":
        sport_score = 0.4
    else:
        sport_score = 0.0

    base_quality = (
        0.35 * deep_score
        + 0.25 * english_score
        + 0.15 * sleep_score
        + 0.15 * sport_score
        + 0.10 * steps_score
    )
    return min(115, int(round(base_quality * 100 + uni_bonus)))


def day_minimum_met(data: dict) -> tuple[bool, dict]:
    english = parse_sheet_number(data.get("Английский_мин"))
    ml = parse_sheet_number(data.get("ML_мин"))
    algos = parse_sheet_number(data.get("Алгосы_мин"))
    uni = parse_sheet_number(data.get("ВУЗ_мин"))
    steps = steps_value(data)
    sleep_hours = parse_sleep_hours(data.get("Сон_часы")) or 0.0
    training = normalize_choice(data.get("Тренировка"))
    reading_pages = parse_sheet_number(data.get("Чтение_стр"))

    any_data = any(
        [
            training,
            english > 0,
            ml > 0,
            algos > 0,
            uni > 0,
            steps > 0,
            sleep_hours > 0,
            reading_pages > 0,
            is_set(data.get("Ккал")),
            parse_sheet_number(data.get("Траты_всего")) > 0,
            is_set(data.get("Вес")),
            is_set(data.get("Настроение")),
            is_set(data.get("Энергия")),
        ]
    )
    min_ok = english >= 30 and max(ml, algos) >= 60 and steps >= 6000 and training != ""
    return min_ok, {
        "english": english,
        "ml": ml,
        "algos": algos,
        "study_total": ml + algos,
        "uni": uni,
        "sleep_hours": sleep_hours,
        "steps": steps,
        "training": training,
        "reading_pages": reading_pages,
        "any_data": any_data,
    }


def day_completion_status(data: dict) -> str:
    min_ok, context = day_minimum_met(data)
    if not context["any_data"]:
        return "empty"
    if min_ok:
        return "full"
    checks = [
        context["english"] >= 30,
        max(context["ml"], context["algos"]) >= 60,
        context["steps"] >= 6000,
        bool(context["training"]),
    ]
    return "partial" if sum(checks) >= 3 else "none"


def compute_missing(data: dict) -> str | None:
    missing: list[str] = []
    training = normalize_choice(data.get("Тренировка"))
    if not training:
        missing.append("Спорт")
    steps = steps_value(data)
    if steps < 6000:
        missing.append("Шаги ≥6k")
    english = parse_sheet_number(data.get("Английский_мин"))
    if english < 30:
        missing.append("Английский ≥30м")
    ml = parse_sheet_number(data.get("ML_мин"))
    algos = parse_sheet_number(data.get("Алгосы_мин"))
    if max(ml, algos) < 60:
        missing.append("ML/Алгосы ≥60м")
    return ", ".join(missing) if missing else None


def mark_set_buttons(buttons: list[tuple[str, str]], current_value: object) -> list[tuple[str, str]]:
    current = "" if current_value is None else str(current_value)
    if current.endswith(".0"):
        current = current[:-2]
    marked: list[tuple[str, str]] = []
    for label, data in buttons:
        if data.startswith("set:"):
            parts = data.split(":", 2)
            value = parts[2] if len(parts) > 2 else ""
            if value == current:
                label = f"✅ {label}"
        marked.append((label, data))
    return marked


def mark_choice_buttons(buttons: list[tuple[str, str]], current_value: object, prefix: str) -> list[tuple[str, str]]:
    current = normalize_choice(current_value)
    marked: list[tuple[str, str]] = []
    for label, data in buttons:
        if data.startswith(prefix):
            value = data.split(":", 1)[1] if ":" in data else ""
            if normalize_choice(value) == current and current:
                label = f"✅ {label}"
        marked.append((label, data))
    return marked


def get_daily_data(context: ContextTypes.DEFAULT_TYPE, date_str: str) -> dict:
    db = get_sheets(context)
    row = db.get_daily_row(date_str)
    if not row:
        return {}

    data: dict[str, object] = {}
    data["Дата"] = date_str
    for db_key, header in DB_TO_HEADER.items():
        data[header] = row.get(db_key)

    habits_done = db.get_habits_done(date_str)
    if habits_done:
        data["Привычки"] = format_habits_value(habits_done)

    # Normalize steps category from raw steps if needed
    if not data.get("Шаги_категория") and data.get("Шаги_кол-во") is not None:
        data["Шаги_категория"] = steps_to_category(parse_sheet_number(data.get("Шаги_кол-во")))

    macros: dict | None = None
    food_tracked = bool(data.get("Еда_учтена"))
    if data.get("Еда_ккал") is not None:
        macros = {
            "kcal": parse_sheet_number(data.get("Еда_ккал")),
            "protein": parse_sheet_number(data.get("Еда_Б")),
            "fat": parse_sheet_number(data.get("Еда_Ж")),
            "carb": parse_sheet_number(data.get("Еда_У")),
        }
        food_tracked = True
    else:
        macros = db.get_daily_macros(date_str)
        if macros:
            food_tracked = True

    data["Еда_учтена"] = 1 if food_tracked else 0
    if macros:
        data["Ккал"] = macros["kcal"]
        data["Белки"] = macros["protein"]
        data["Жиры"] = macros["fat"]
        data["Угли"] = macros["carb"]
    else:
        data["Ккал"] = None
        data["Белки"] = None
        data["Жиры"] = None
        data["Угли"] = None
    data["_macros"] = macros

    expenses = db.get_expense_totals(date_str)
    data["_expenses"] = expenses
    data["Траты_всего"] = expenses.get("total", 0.0)
    data["Траты_еда"] = expenses.get("Еда", 0.0)
    data["Траты_одежда"] = expenses.get("Одежда", 0.0)
    data["Траты_бытовуха"] = expenses.get("Бытовуха", 0.0)
    data["Траты_гульки"] = expenses.get("Гульки", 0.0)
    data["Траты_здоровье"] = expenses.get("Здоровье", 0.0)
    data["Траты_другое"] = expenses.get("Другое", 0.0)

    data["_sleep_start"] = db.get_state(STATE_SLEEP_START)
    data["_sleep_start_day"] = db.get_state(STATE_SLEEP_START_DAY)
    data["_active_day"] = db.get_state(STATE_ACTIVE_DAY)

    quality = compute_quality(data)
    data["Качество_дня"] = quality if quality is not None else ""
    data["Коэффициент_дня"] = data["Качество_дня"]
    missing = compute_missing(data)
    data["Не_заполнено"] = missing or ""
    return data


def normalize_choice(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def parse_habits_value(value: object) -> list[str]:
    if value in (None, ""):
        return []
    text = str(value).replace(",", ";")
    return [part.strip() for part in text.split(";") if part.strip()]


def format_habits_value(items: list[str]) -> str:
    return "; ".join(items)


FIELD_HEADERS = {
    "training": "Тренировка",
    "cardio": "Кардио_мин",
    "steps": "Шаги_категория",
    "english": "Английский_мин",
    "ml": "ML_мин",
    "algos": "Алгосы_мин",
    "uni": "ВУЗ_мин",
    "code_mode": "Код_режим",
    "code_topic": "Код_тема",
    "reading": "Чтение_стр",
    "rest_time": "Отдых_время",
    "rest_type": "Отдых_тип",
    "sleep_bed": "Сон_отбой",
    "sleep_hours": "Сон_часы",
    "nap": "Сон_дневной",
    "sleep_regime": "Режим",
    "productivity": "Продуктивность",
    "mood": "Настроение",
    "energy": "Энергия",
}

FIELD_LABELS = {
    "training": "Тренировка",
    "cardio": "Кардио",
    "steps": "Шаги",
    "english": "Английский",
    "ml": "ML",
    "algos": "Алгосы",
    "uni": "ВУЗ",
    "code_mode": "Код (режим)",
    "code_topic": "Код (тема)",
    "reading": "Чтение",
    "rest_time": "Отдых (время)",
    "rest_type": "Отдых (тип)",
    "sleep_bed": "Сон (отбой)",
    "sleep_hours": "Сон (часы)",
    "nap": "Дневной сон",
    "sleep_regime": "Режим",
    "productivity": "Продуктивность",
    "mood": "Настроение",
    "energy": "Энергия",
}


def build_sport_menu(data: dict) -> list[tuple[str, str]]:
    training = data.get("Тренировка")
    training_display = display_training(training) if training in {"Ноги", "Верх", "Низ", "Фулл"} else None
    training_selected = training_display is not None
    training_label = "Тренировка"
    if training_display:
        training_label = f"Тренировка: {training_display}"

    rest_selected = training == "Отдых"
    skip_selected = training == "Пропустил"

    cardio = data.get("Кардио_мин")
    cardio_label = "Кардио"
    if cardio not in (None, ""):
        cardio_label = f"Кардио: {cardio}м"

    buttons = [
        (f"✅ {training_label}" if training_selected else training_label, "sport:training"),
        ("✅ Отдых" if rest_selected else "Отдых", "sport:rest"),
        ("✅ Пропуск" if skip_selected else "Пропуск", "sport:skip"),
        (f"✅ {cardio_label}" if cardio not in (None, "") else cardio_label, "sport:cardio"),
    ]
    return buttons


def build_study_menu(data: dict) -> list[tuple[str, str]]:
    english = data.get("Английский_мин")
    english_label = "Английский" if english in (None, "") else f"Английский: {english}м"

    ml = data.get("ML_мин")
    ml_label = "ML" if ml in (None, "") else f"ML: {ml}м"

    algos = data.get("Алгосы_мин")
    algos_label = "Алгосы" if algos in (None, "") else f"Алгосы: {algos}м"

    uni = data.get("ВУЗ_мин")
    uni_label = "ВУЗ" if uni in (None, "") else f"ВУЗ: {uni}м"

    reading = data.get("Чтение_стр")
    reading_label = format_reading_label(reading)

    return [
        (f"✅ {english_label}" if english not in (None, "") else english_label, "study:english"),
        (f"✅ {ml_label}" if ml not in (None, "") else ml_label, "study:ml"),
        (f"✅ {algos_label}" if algos not in (None, "") else algos_label, "study:algos"),
        (f"✅ {uni_label}" if uni not in (None, "") else uni_label, "study:uni"),
        (f"✅ {reading_label}" if reading_is_set(reading) else reading_label, "study:reading"),
    ]


def build_leisure_menu(data: dict) -> list[tuple[str, str]]:
    rest_time = data.get("Отдых_время")
    rest_label = "Отдых" if rest_time in (None, "") else f"Отдых: {rest_time}"

    productivity = data.get("Продуктивность")
    prod_label = "Продуктивность" if productivity in (None, "") else f"Продуктивность: {productivity}%"

    shots = data.get("Стрельнул_раз")
    shots_value = int(parse_sheet_number(shots)) if shots not in (None, "") else 0
    shots_label = f"Стрельнул: {shots_value}"

    sleep_hours = data.get("Сон_часы")
    sleep_label = "Сон (вручную)"
    if sleep_hours not in (None, ""):
        sleep_label = f"Сон: {sleep_hours}ч (ред.)"

    nap_hours = data.get("Сон_дневной")
    nap_label = "Дневной сон"
    if nap_hours not in (None, ""):
        nap_label = f"Дневной сон: {nap_hours}ч"

    anti_count = data.get("_anti_count")
    anti_label = "Анти‑прокраст."
    if anti_count:
        anti_label = f"Анти‑прокраст.: {anti_count}"

    spend_total = parse_sheet_number(data.get("Траты_всего"))
    spend_label = "Траты" if spend_total <= 0 else f"Траты: {fmt_money(spend_total)} ₽"

    return [
        (f"✅ {rest_label}" if rest_time not in (None, "") else rest_label, "leisure:rest"),
        (f"✅ {prod_label}" if productivity not in (None, "") else prod_label, "leisure:productivity"),
        (f"✅ {spend_label}" if spend_total > 0 else spend_label, "leisure:expenses"),
        (shots_label, "leisure:shots"),
        (sleep_label, "leisure:sleep_manual"),
        (f"✅ {nap_label}" if nap_hours not in (None, "") else nap_label, "leisure:nap"),
        (f"✅ {anti_label}" if anti_count else anti_label, "leisure:anti"),
    ]


def build_morale_menu(data: dict) -> list[tuple[str, str]]:
    mood = data.get("Настроение")
    energy = data.get("Энергия")
    return [
        (f"✅ Настроение: {mood}" if mood not in (None, "") else "Настроение", "morale:mood"),
        (f"✅ Энергия: {energy}" if energy not in (None, "") else "Энергия", "morale:energy"),
        ("О чем жалею", "morale:regret"),
        ("Отзыв о дне", "morale:review"),
    ]


def build_code_buttons(current_mode: object) -> list[tuple[str, str]]:
    buttons = mark_choice_buttons(CODE_MODE_OPTIONS, current_mode, "code_mode:")
    buttons.append(("↩️ Удалить последнюю", "code:undo"))
    buttons.append(("🗑 Очистить код за сегодня", "code:clear"))
    return buttons


async def build_code_menu(context: ContextTypes.DEFAULT_TYPE, date_str: str) -> tuple[str, list[tuple[str, str]]]:
    sheets = get_sheets(context)
    sessions = sheets.get_sessions(date_str, category="Код")
    current_mode = context.user_data.get("code_mode")

    lines = ["💻 Код за сегодня"]
    if not sessions:
        lines.append("Пока нет записей.")
    else:
        labels = [s.get("subcategory") for s in sessions if s.get("subcategory")]
        max_items = 6
        for label in labels[:max_items]:
            lines.append(f"• {label}")
        if len(labels) > max_items:
            lines.append(f"… ещё {len(labels) - max_items}")

    if current_mode:
        lines.append("")
        lines.append(f"Выбран режим: {current_mode}")

    return "\n".join(lines), build_code_buttons(current_mode)


async def build_habits_menu(context: ContextTypes.DEFAULT_TYPE, date_str: str) -> tuple[str, list[tuple[str, str]]]:
    sheets = get_sheets(context)
    habits = sheets.get_habits()
    daily = get_daily_data(context, date_str)
    completed = set(parse_habits_value(daily.get("Привычки")))

    total = len(habits)
    done = sum(1 for habit in habits if habit in completed)
    if total:
        header = f"🧠 Привычки сегодня: {done}/{total}"
    else:
        header = "🧠 Привычки: пока нет списка"

    buttons: list[tuple[str, str]] = []
    context.user_data["habit_list"] = habits
    max_items = 10
    for idx, habit in enumerate(habits[:max_items]):
        status = "✅" if habit in completed else "⬜"
        buttons.append((f"{status} {habit}", f"habit:toggle:{idx}"))
    if len(habits) > max_items:
        header = f"{header}\n… ещё {len(habits) - max_items} привычек не показаны"

    buttons.append(("➕ Добавить привычку", "habit:add"))
    if completed:
        buttons.append(("🧹 Сбросить отметки", "habit:clear"))

    return header, buttons


def sync_code_fields(sheets: Database, date_str: str) -> None:
    sessions = sheets.get_sessions(date_str, category="Код")
    if not sessions:
        sheets.update_daily_fields(date_str, {COLUMN_MAP["code_mode"]: "", COLUMN_MAP["code_topic"]: ""})
        return
    last = sessions[-1].get("subcategory") or ""
    if "/" in last:
        mode, topic = last.split("/", 1)
    else:
        mode, topic = last, ""
    sheets.update_daily_fields(date_str, {COLUMN_MAP["code_mode"]: mode, COLUMN_MAP["code_topic"]: topic})


async def build_anti_menu(context: ContextTypes.DEFAULT_TYPE, date_str: str) -> tuple[str, list[tuple[str, str]]]:
    sheets = get_sheets(context)
    sessions = sheets.get_sessions(date_str, category="Анти")
    counts: dict[str, int] = {}
    for item in sessions:
        reason = item.get("subcategory") or ""
        if not reason:
            continue
        counts[reason] = counts.get(reason, 0) + 1

    lines = ["🧯 Анти‑прокрастинация"]
    if counts:
        summary = ", ".join(f"{k}×{v}" for k, v in list(counts.items())[:6])
        lines.append(f"Сегодня: {summary}")
    else:
        lines.append("Сегодня пока нет отметок.")

    buttons: list[tuple[str, str]] = []
    for label, data in PROCRASTINATION_OPTIONS:
        count = counts.get(label)
        display = f"{label} ({count})" if count else label
        buttons.append((display, data))
    buttons.append(("✍️ Другое", "anti:custom"))
    buttons.append(("↩️ Удалить последнюю", "anti:undo"))
    if counts:
        buttons.append(("🗑 Очистить сегодня", "anti:clear"))

    return "\n".join(lines), buttons


async def build_expense_menu(context: ContextTypes.DEFAULT_TYPE, date_str: str) -> tuple[str, list[tuple[str, str]]]:
    db = get_sheets(context)
    totals = db.get_expense_totals(date_str)
    total = float(totals.get("total", 0.0))

    lines = [
        "💸 Траты за день",
        f"Итого: {fmt_money(total)} ₽",
    ]
    non_zero = [(label, float(totals.get(label, 0.0))) for label in EXPENSE_CATEGORY_LABELS.values() if totals.get(label)]
    if non_zero:
        details = " · ".join(f"{label.lower()} {fmt_money(amount)}" for label, amount in non_zero[:4])
        lines.append(details)
        if len(non_zero) > 4:
            lines.append(f"… ещё {len(non_zero) - 4} катег.")

    buttons: list[tuple[str, str]] = []
    for label, callback in EXPENSE_OPTIONS:
        amount = float(totals.get(label, 0.0))
        btn = f"{label}: {fmt_money(amount)} ₽" if amount > 0 else label
        buttons.append((btn, callback))

    buttons.append(("↩️ Удалить последнюю", "expense:undo"))
    if total > 0:
        buttons.append(("🗑 Очистить за день", "expense:clear"))
    return "\n".join(lines), buttons


def build_code_label(sessions: list[dict]) -> tuple[str, bool]:
    if not sessions:
        return ("Код", False)
    labels = [s.get("subcategory") for s in sessions if s.get("subcategory")]
    preview = ", ".join(labels[:2])
    if len(labels) > 2:
        preview = f"{preview} +{len(labels) - 2}"
    label = f"Код: {preview}" if preview else f"Код: {len(labels)}"
    return (label, True)


async def show_study_menu(query, context: ContextTypes.DEFAULT_TYPE, date_str: str) -> None:
    daily = get_daily_data(context, date_str)
    await show_menu(query, "Учеба:", build_study_menu(daily))


def recommend_portions(
    current: dict,
    target_mid: dict,
    portions: list[dict],
    *,
    eaten_products: set[str],
    max_items: int = 3,
) -> list[dict]:
    weights = {"kcal": 1.0, "protein": 1.3, "fat": 0.8, "carb": 1.0}

    def deficit(vec):
        return {
            key: max(0.0, target_mid[key] - vec.get(key, 0.0))
            for key in target_mid
        }

    base_def = deficit(current)
    base_score = sum((base_def[k] ** 2) * weights[k] for k in base_def)

    scored = []
    for portion in portions:
        new_vec = {
            key: current.get(key, 0.0) + portion["macros"].get(key, 0.0)
            for key in target_mid
        }
        new_def = deficit(new_vec)
        new_score = sum((new_def[k] ** 2) * weights[k] for k in new_def)
        improvement = base_score - new_score
        if improvement <= 0:
            continue
        penalty = 0.6 if portion["product"] in eaten_products else 1.0
        scored.append((improvement * penalty, portion))

    scored.sort(key=lambda x: x[0], reverse=True)
    return [item for _, item in scored[:max_items]]


def build_plan(
    current: dict,
    target_mid: dict,
    portions: list[dict],
    *,
    eaten_products: set[str],
    max_steps: int = 4,
) -> list[dict]:
    plan: list[dict] = []
    temp = current.copy()
    used_products: set[str] = set()
    for _ in range(max_steps):
        candidates = recommend_portions(
            temp,
            target_mid,
            portions,
            eaten_products=eaten_products.union(used_products),
            max_items=1,
        )
        if not candidates:
            break
        choice = candidates[0]
        plan.append(choice)
        used_products.add(choice["product"])
        for key in target_mid:
            temp[key] = temp.get(key, 0.0) + choice["macros"].get(key, 0.0)
    return plan


def menu_config(menu_key: str, data: dict) -> tuple[str, list[tuple[str, str]], str, int]:
    if menu_key == "sport":
        return ("Спорт:", build_sport_menu(data), "menu:main", 2)
    if menu_key == "study":
        return ("Учеба:", build_study_menu(data), "menu:main", 2)
    if menu_key == "leisure":
        return ("Досуг:", build_leisure_menu(data), "menu:main", 2)
    if menu_key == "morale":
        return ("Моралька:", build_morale_menu(data), "menu:main", 2)
    if menu_key == "training":
        return ("Тренировка:", mark_set_buttons(TRAINING_OPTIONS, data.get("Тренировка")), "menu:sport", 2)
    if menu_key == "cardio":
        return ("Кардио (мин):", mark_set_buttons(CARDIO_OPTIONS, data.get("Кардио_мин")), "menu:sport", 3)
    if menu_key == "steps":
        return ("Шаги:", mark_set_buttons(STEPS_OPTIONS, data.get("Шаги_категория")), "menu:sport", 2)
    if menu_key == "english":
        return ("Английский:", mark_set_buttons(ENGLISH_OPTIONS, data.get("Английский_мин")), "menu:study", 3)
    if menu_key == "ml":
        return ("ML:", mark_set_buttons(ML_OPTIONS, data.get("ML_мин")), "menu:study", 3)
    if menu_key == "algos":
        return ("Алгосы:", mark_set_buttons(ALGOS_OPTIONS, data.get("Алгосы_мин")), "menu:study", 3)
    if menu_key == "uni":
        return ("ВУЗ:", mark_set_buttons(UNI_OPTIONS, data.get("ВУЗ_мин")), "menu:study", 3)
    if menu_key == "code_mode":
        return ("Код: режим", mark_set_buttons(CODE_MODE_OPTIONS, data.get("Код_режим")), "menu:study", 1)
    if menu_key == "code_topic":
        return ("Код: тема", mark_set_buttons(CODE_TOPIC_OPTIONS, data.get("Код_тема")), "menu:study", 2)
    if menu_key == "reading":
        return ("Чтение:", mark_set_buttons(READING_OPTIONS, data.get("Чтение_стр")), "menu:study", 4)
    if menu_key == "rest_time":
        return ("Отдых: время", mark_set_buttons(REST_TIME_OPTIONS, data.get("Отдых_время")), "menu:leisure", 2)
    if menu_key == "rest_type":
        return ("Отдых: тип", mark_set_buttons(REST_TYPE_OPTIONS, data.get("Отдых_тип")), "menu:leisure", 2)
    if menu_key == "sleep_bed":
        return ("Сон: во сколько заснул?", mark_set_buttons(SLEEP_BEDTIME_OPTIONS, data.get("Сон_отбой")), "menu:leisure", 3)
    if menu_key == "sleep_hours":
        return ("Сон: сколько часов?", mark_set_buttons(SLEEP_HOURS_OPTIONS, data.get("Сон_часы")), "menu:leisure", 3)
    if menu_key == "sleep_regime":
        return ("Сон: режим", mark_set_buttons(SLEEP_REGIME_OPTIONS, data.get("Режим")), "menu:leisure", 2)
    if menu_key == "nap":
        return ("Дневной сон:", mark_set_buttons(NAP_OPTIONS, data.get("Сон_дневной")), "menu:leisure", 2)
    if menu_key == "productivity":
        return ("Продуктивность:", mark_set_buttons(PRODUCTIVITY_OPTIONS, data.get("Продуктивность")), "menu:leisure", 3)
    if menu_key == "mood":
        return ("Настроение:", mark_set_buttons(MOOD_OPTIONS, data.get("Настроение")), "menu:morale", 2)
    if menu_key == "energy":
        return ("Энергия:", mark_set_buttons(ENERGY_OPTIONS, data.get("Энергия")), "menu:morale", 2)
    return ("Главное меню:", MAIN_MENU, "menu:main", 2)


async def confirm_override(
    context: ContextTypes.DEFAULT_TYPE,
    query,
    *,
    field_key: str,
    current_value: object,
    new_value: object,
    return_menu: str,
    next_menu: str | None = None,
) -> None:
    label = FIELD_LABELS.get(field_key, field_key)
    current_display = display_training(current_value) if field_key == "training" else fmt_value(current_value)
    new_display = display_training(new_value) if field_key == "training" else fmt_value(new_value)
    context.user_data["pending_set"] = {
        "field_key": field_key,
        "value": new_value,
        "return_menu": return_menu,
        "next_menu": next_menu,
    }
    buttons = [("✅ Да", "confirm:yes"), ("↩️ Нет", "confirm:no")]
    back_cb = return_menu if return_menu.startswith("menu:") else f"menu:{return_menu}"
    await query.answer()
    await query.edit_message_text(
        f"⚠️ {label}\nСейчас: {current_display}\nЗаменить на: {new_display}?",
        reply_markup=build_keyboard(buttons, cols=2, back=("⬅️ Назад", back_cb)),
    )


async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_authorized(context, update.effective_user.id if update.effective_user else None):
        return
    query = update.callback_query
    data = query.data
    cfg = context.application.bot_data["config"]
    sheets = get_sheets(context)
    date_str = get_view_date(context)
    sheets.ensure_daily_row(date_str)

    if data.startswith("confirm:"):
        await query.answer()
        pending = context.user_data.get("pending_set")
        if not pending:
            await query.edit_message_text("Главное меню:", reply_markup=build_keyboard(MAIN_MENU, cols=2))
            return
        if data == "confirm:yes":
            field_key = pending["field_key"]
            value = pending["value"]
            next_menu = pending.get("next_menu")
            return_menu = pending.get("return_menu", "menu:main")
            sheets.update_daily_fields(date_str, {COLUMN_MAP[field_key]: value})
            context.user_data.pop("pending_set", None)
            daily = get_daily_data(context, date_str)
            menu_key = next_menu or return_menu
            if menu_key == "study":
                await show_study_menu(query, context, date_str)
                return
            title, buttons, back_to, cols = menu_config(menu_key, daily)
            await show_menu(query, title, buttons, back_to=back_to, cols=cols)
            return
        if data == "confirm:no":
            return_menu = pending.get("return_menu", "menu:main")
            context.user_data.pop("pending_set", None)
            daily = get_daily_data(context, date_str)
            if return_menu == "study":
                await show_study_menu(query, context, date_str)
                return
            title, buttons, back_to, cols = menu_config(return_menu, daily)
            await show_menu(query, title, buttons, back_to=back_to, cols=cols)
            return

    if data == "menu:main":
        await query.answer()
        db = get_sheets(context)
        if query.message is not None:
            db.set_state(summary_state_key(query.message.chat_id), str(query.message.message_id))
        await safe_render_summary(context, query.message.chat_id, date_str)
        return
    if data == "menu:refresh":
        await query.answer()
        db = get_sheets(context)
        if query.message is not None:
            db.set_state(summary_state_key(query.message.chat_id), str(query.message.message_id))
        await safe_render_summary(context, query.message.chat_id, date_str)
        return
    if data.startswith("stats:"):
        await query.answer()
        if data == "stats:back":
            await safe_render_summary(context, query.message.chat_id, date_str)
            return
        period = data.split(":", 1)[1]
        await render_stats(context, query.message.chat_id, period)
        return
    if data == "menu:date":
        await query.answer()
        await send_or_edit_prompt(
            context,
            query.message.chat_id,
            "Выбери дату просмотра:",
            build_keyboard(
                [("Сегодня", "date:today"), ("Вчера", "date:yesterday"), ("Ввести дату", "date:pick")],
                cols=2,
                back=("⬅️ Назад", "menu:main"),
            ),
        )
        return
    if data == "date:today":
        await query.answer()
        set_view_date(context, get_active_date(context))
        await clear_prompt(context, query.message.chat_id)
        await safe_render_summary(context, query.message.chat_id, get_view_date(context))
        return
    if data == "date:yesterday":
        await query.answer()
        yday = (get_now(cfg.timezone).date() - timedelta(days=1)).isoformat()
        set_view_date(context, yday)
        await clear_prompt(context, query.message.chat_id)
        await safe_render_summary(context, query.message.chat_id, get_view_date(context))
        return
    if data == "date:pick":
        await query.answer()
        context.user_data["expect"] = "view_date"
        await send_or_edit_prompt(
            context,
            query.message.chat_id,
            "Введи дату в формате YYYY-MM-DD (например 2026-02-12)",
            build_keyboard([("⬅️ Назад", "menu:main")], cols=1),
        )
        return
    if data == "menu:sport":
        daily = get_daily_data(context, date_str)
        await show_menu(query, "Спорт:", build_sport_menu(daily))
        return
    if data == "menu:study":
        await show_study_menu(query, context, date_str)
        return
    if data == "menu:leisure":
        daily = get_daily_data(context, date_str)
        anti_sessions = sheets.get_sessions(date_str, category="Анти")
        if anti_sessions:
            daily["_anti_count"] = len(anti_sessions)
        await show_menu(query, "Досуг:", build_leisure_menu(daily))
        return
    if data == "menu:food":
        await query.answer()
        summary = await build_food_summary(context, date_str)
        await query.edit_message_text(
            summary,
            reply_markup=build_keyboard(FOOD_MENU, cols=2, back=("⬅️ Назад", "menu:main")),
        )
        return
    if data == "menu:morale":
        daily = get_daily_data(context, date_str)
        await show_menu(query, "Моралька:", build_morale_menu(daily))
        return
    if data == "menu:habits":
        await query.answer()
        text, buttons = await build_habits_menu(context, date_str)
        await query.edit_message_text(
            text,
            reply_markup=build_keyboard(buttons, cols=1, back=("⬅️ Назад", "menu:main")),
        )
        return

    if data.startswith("habit:toggle:"):
        await query.answer()
        try:
            idx = int(data.split(":")[2])
        except (IndexError, ValueError):
            return
        habits = context.user_data.get("habit_list") or sheets.get_habits()
        if idx < 0 or idx >= len(habits):
            return
        habit = habits[idx]
        daily = get_daily_data(context, date_str)
        completed = parse_habits_value(daily.get("Привычки"))
        if habit in completed:
            completed = [h for h in completed if h != habit]
            sheets.set_habit_done(date_str, habit, False)
        else:
            completed.append(habit)
            sheets.set_habit_done(date_str, habit, True)
        sheets.update_daily_fields(date_str, {COLUMN_MAP["habits"]: format_habits_value(completed)})
        text, buttons = await build_habits_menu(context, date_str)
        await query.edit_message_text(
            text,
            reply_markup=build_keyboard(buttons, cols=1, back=("⬅️ Назад", "menu:main")),
        )
        return

    if data == "habit:add":
        await query.answer()
        context.user_data["expect"] = "habit_add"
        await send_or_edit_prompt(
            context,
            query.message.chat_id,
            "Введи название новой привычки (например, \"Зарядка\"):",
        )
        return

    if data == "habit:clear":
        await query.answer()
        await query.edit_message_text(
            "Сбросить все отметки привычек за сегодня?",
            reply_markup=build_keyboard([("✅ Да", "habit_clear:yes"), ("↩️ Нет", "habit_clear:no")], cols=2, back=("⬅️ Назад", "menu:habits")),
        )
        return

    if data.startswith("habit_clear:"):
        await query.answer()
        if data == "habit_clear:yes":
            sheets.clear_habits_for_date(date_str)
            sheets.update_daily_fields(date_str, {COLUMN_MAP["habits"]: ""})
        text, buttons = await build_habits_menu(context, date_str)
        await query.edit_message_text(
            text,
            reply_markup=build_keyboard(buttons, cols=1, back=("⬅️ Назад", "menu:main")),
        )
        return

    if data.startswith("anti:"):
        await query.answer()
        reason = data.split(":", 1)[1] if ":" in data else ""
        if reason == "custom":
            context.user_data["expect"] = "anti_custom"
            await send_or_edit_prompt(
                context,
                query.message.chat_id,
                "Напиши причину прокрастинации (коротко):",
            )
            return
        if reason == "undo":
            removed = sheets.delete_last_session(date_str, category="Анти")
            text, buttons = await build_anti_menu(context, date_str)
            prefix = "↩️ Удалил последнюю.\n\n" if removed else "Нет записей для удаления.\n\n"
            await query.edit_message_text(
                f"{prefix}{text}",
                reply_markup=build_keyboard(buttons, cols=2, back=("⬅️ Назад", "menu:leisure")),
            )
            return
        if reason == "clear":
            sheets.clear_sessions(date_str, category="Анти")
            text, buttons = await build_anti_menu(context, date_str)
            await query.edit_message_text(
                f"🗑 Очистил записи.\n\n{text}",
                reply_markup=build_keyboard(buttons, cols=2, back=("⬅️ Назад", "menu:leisure")),
            )
            return
        # regular reason
        sheets.add_session(date_str, time_str(cfg.timezone), "Анти", reason, 0, "")
        text, buttons = await build_anti_menu(context, date_str)
        await query.edit_message_text(
            text,
            reply_markup=build_keyboard(buttons, cols=2, back=("⬅️ Назад", "menu:leisure")),
        )
        return

    if data == "sport:training":
        daily = get_daily_data(context, date_str)
        current = daily.get("Тренировка")
        await show_menu(query, "Тренировка:", mark_set_buttons(TRAINING_OPTIONS, current), back_to="menu:sport", cols=2)
        return
    if data == "sport:rest":
        daily = get_daily_data(context, date_str)
        current = daily.get("Тренировка")
        new_value = "Отдых"
        if normalize_choice(current) and normalize_choice(current) != normalize_choice(new_value):
            await confirm_override(
                context,
                query,
                field_key="training",
                current_value=current,
                new_value=new_value,
                return_menu="sport",
            )
            return
        sheets.update_daily_fields(date_str, {COLUMN_MAP["training"]: new_value})
        daily = get_daily_data(context, date_str)
        await show_menu(query, "Спорт:", build_sport_menu(daily))
        return
    if data == "sport:skip":
        daily = get_daily_data(context, date_str)
        current = daily.get("Тренировка")
        new_value = "Пропустил"
        if normalize_choice(current) and normalize_choice(current) != normalize_choice(new_value):
            await confirm_override(
                context,
                query,
                field_key="training",
                current_value=current,
                new_value=new_value,
                return_menu="sport",
            )
            return
        sheets.update_daily_fields(date_str, {COLUMN_MAP["training"]: new_value})
        daily = get_daily_data(context, date_str)
        await show_menu(query, "Спорт:", build_sport_menu(daily))
        return
    if data == "sport:cardio":
        daily = get_daily_data(context, date_str)
        current = daily.get("Кардио_мин")
        await show_menu(query, "Кардио (мин):", mark_set_buttons(CARDIO_OPTIONS, current), back_to="menu:sport", cols=3)
        return
    if data == "sport:steps":
        daily = get_daily_data(context, date_str)
        current = daily.get("Шаги_категория")
        await show_menu(query, "Шаги:", mark_set_buttons(STEPS_OPTIONS, current), back_to="menu:sport", cols=2)
        return

    if data == "study:english":
        daily = get_daily_data(context, date_str)
        current = daily.get("Английский_мин")
        await show_menu(query, "Английский:", mark_set_buttons(ENGLISH_OPTIONS, current), back_to="menu:study", cols=3)
        return
    if data == "study:ml":
        daily = get_daily_data(context, date_str)
        current = daily.get("ML_мин")
        await show_menu(query, "ML:", mark_set_buttons(ML_OPTIONS, current), back_to="menu:study", cols=3)
        return
    if data == "study:algos":
        daily = get_daily_data(context, date_str)
        current = daily.get("Алгосы_мин")
        await show_menu(query, "Алгосы:", mark_set_buttons(ALGOS_OPTIONS, current), back_to="menu:study", cols=3)
        return
    if data == "study:uni":
        daily = get_daily_data(context, date_str)
        current = daily.get("ВУЗ_мин")
        await show_menu(query, "ВУЗ:", mark_set_buttons(UNI_OPTIONS, current), back_to="menu:study", cols=3)
        return
    if data == "study:code":
        await query.answer()
        text, buttons = await build_code_menu(context, date_str)
        await query.edit_message_text(
            text,
            reply_markup=build_keyboard(buttons, cols=1, back=("⬅️ Назад", "menu:study")),
        )
        return
    if data == "study:reading":
        daily = get_daily_data(context, date_str)
        current = daily.get("Чтение_стр")
        await show_menu(query, "Чтение:", mark_set_buttons(READING_OPTIONS, current), back_to="menu:study", cols=4)
        return

    if data.startswith("code_mode:"):
        mode = data.split(":", 1)[1]
        context.user_data["code_mode"] = mode
        await query.answer()
        await query.edit_message_text(
            f"💻 Режим: {mode}\nВыбери тему:",
            reply_markup=build_keyboard(mark_choice_buttons(CODE_TOPIC_OPTIONS, None, "code_topic:"), cols=2, back=("⬅️ Назад", "study:code")),
        )
        return

    if data.startswith("code_topic:"):
        topic = data.split(":", 1)[1]
        mode = context.user_data.get("code_mode")
        if not mode:
            await query.answer()
            await query.edit_message_text(
                "Сначала выбери режим:",
                reply_markup=build_keyboard(mark_choice_buttons(CODE_MODE_OPTIONS, None, "code_mode:"), cols=2, back=("⬅️ Назад", "menu:study")),
            )
            return
        sheets.add_session(date_str, time_str(cfg.timezone), "Код", f"{mode}/{topic}", 0, "")
        sync_code_fields(sheets, date_str)
        context.user_data.pop("code_mode", None)
        text, buttons = await build_code_menu(context, date_str)
        await query.answer()
        await query.edit_message_text(
            f"✅ Добавил: {mode}/{topic}\n\n{text}",
            reply_markup=build_keyboard(buttons, cols=1, back=("⬅️ Назад", "menu:study")),
        )
        return

    if data == "code:undo":
        removed = sheets.delete_last_session(date_str, category="Код")
        if removed:
            sync_code_fields(sheets, date_str)
        text, buttons = await build_code_menu(context, date_str)
        await query.answer()
        prefix = "↩️ Удалил последнюю запись.\n\n" if removed else "Нет записей для удаления.\n\n"
        await query.edit_message_text(
            f"{prefix}{text}",
            reply_markup=build_keyboard(buttons, cols=1, back=("⬅️ Назад", "menu:study")),
        )
        return

    if data == "code:clear":
        await query.answer()
        context.user_data["pending_code_clear"] = True
        await query.edit_message_text(
            "Удалить все записи кода за сегодня?",
            reply_markup=build_keyboard([("✅ Да", "code_clear:yes"), ("↩️ Нет", "code_clear:no")], cols=2, back=("⬅️ Назад", "menu:study")),
        )
        return

    if data.startswith("code_clear:"):
        await query.answer()
        if data == "code_clear:yes":
            sheets.clear_sessions(date_str, category="Код")
            sync_code_fields(sheets, date_str)
            context.user_data.pop("pending_code_clear", None)
            text, buttons = await build_code_menu(context, date_str)
            await query.edit_message_text(
                f"🗑 Очистил записи.\n\n{text}",
                reply_markup=build_keyboard(buttons, cols=1, back=("⬅️ Назад", "menu:study")),
            )
            return
        if data == "code_clear:no":
            context.user_data.pop("pending_code_clear", None)
            text, buttons = await build_code_menu(context, date_str)
            await query.edit_message_text(
                text,
                reply_markup=build_keyboard(buttons, cols=1, back=("⬅️ Назад", "menu:study")),
            )
            return

    if data == "leisure:rest":
        daily = get_daily_data(context, date_str)
        current = daily.get("Отдых_время")
        await show_menu(query, "Отдых: время", mark_set_buttons(REST_TIME_OPTIONS, current), back_to="menu:leisure", cols=2)
        return
    if data == "leisure:nap":
        daily = get_daily_data(context, date_str)
        current = daily.get("Сон_дневной")
        await show_menu(query, "Дневной сон:", mark_set_buttons(NAP_OPTIONS, current), back_to="menu:leisure", cols=2)
        return
    if data == "leisure:nap_custom":
        await query.answer()
        context.user_data["expect"] = "nap_hours"
        await send_or_edit_prompt(
            context,
            query.message.chat_id,
            "Дневной сон: сколько часов? (например, 1.5)",
            build_keyboard([("⬅️ Назад", "menu:leisure")], cols=1),
        )
        return
    if data == "leisure:shots":
        daily = get_daily_data(context, date_str)
        count = int(parse_sheet_number(daily.get("Стрельнул_раз")))
        await query.answer()
        await query.edit_message_text(
            f"Стрельнул сегодня: {count}",
            reply_markup=build_shots_keyboard(count),
        )
        return
    if data == "leisure:expenses":
        await query.answer()
        text, buttons = await build_expense_menu(context, date_str)
        await query.edit_message_text(
            text,
            reply_markup=build_keyboard(buttons, cols=2, back=("⬅️ Назад", "menu:leisure")),
        )
        return
    if data.startswith("expense:add:"):
        await query.answer()
        category_key = data.split(":", 2)[2]
        category_label = EXPENSE_CATEGORY_LABELS.get(category_key)
        if not category_label:
            return
        context.user_data["expense_category"] = category_label
        context.user_data["expect"] = "expense_amount"
        await send_or_edit_prompt(
            context,
            query.message.chat_id,
            f"💸 {category_label}: введи сумму (например 1450 или 320.5)",
            build_keyboard([("⬅️ Назад", "leisure:expenses")], cols=1),
        )
        return
    if data == "expense:undo":
        await query.answer()
        removed = sheets.delete_last_expense(date_str)
        text, buttons = await build_expense_menu(context, date_str)
        prefix = "↩️ Удалил последнюю трату.\n\n" if removed else "Нет трат для удаления.\n\n"
        await query.edit_message_text(
            f"{prefix}{text}",
            reply_markup=build_keyboard(buttons, cols=2, back=("⬅️ Назад", "menu:leisure")),
        )
        return
    if data == "expense:clear":
        await query.answer()
        await query.edit_message_text(
            "Очистить все траты за выбранный день?",
            reply_markup=build_keyboard(
                [("✅ Да", "expense_clear:yes"), ("↩️ Нет", "expense_clear:no")],
                cols=2,
                back=("⬅️ Назад", "leisure:expenses"),
            ),
        )
        return
    if data.startswith("expense_clear:"):
        await query.answer()
        if data == "expense_clear:yes":
            sheets.clear_expenses(date_str)
        text, buttons = await build_expense_menu(context, date_str)
        await query.edit_message_text(
            text,
            reply_markup=build_keyboard(buttons, cols=2, back=("⬅️ Назад", "menu:leisure")),
        )
        return
    if data == "leisure:sleep_manual":
        await query.answer()
        context.user_data["expect"] = "sleep_bed_manual"
        await send_or_edit_prompt(
            context,
            query.message.chat_id,
            "Во сколько лег спать? (HH:MM)",
            build_keyboard([("⬅️ Назад", "menu:leisure")], cols=1),
        )
        return
    if data == "sleep:edit":
        await query.answer()
        context.user_data["expect"] = "sleep_bed_edit"
        await send_or_edit_prompt(
            context,
            query.message.chat_id,
            "Во сколько реально заснул? (HH:MM)",
            build_keyboard([("⬅️ Назад", "menu:main")], cols=1),
        )
        return
    if data == "sleep:cancel":
        await query.answer()
        await send_or_edit_prompt(
            context,
            query.message.chat_id,
            "Отменить сон?",
            build_keyboard(
                [("✅ Отменить сон", "sleep:cancel_yes"), ("☀️ Я проснулся", "sleep:cancel_wake")],
                cols=2,
                back=("⬅️ Назад", "menu:main"),
            ),
        )
        return
    if data == "sleep:cancel_yes":
        await query.answer()
        sheets.set_state(STATE_SLEEP_START, None)
        sheets.set_state(STATE_SLEEP_START_DAY, None)
        sheets.set_state(STATE_SLEEP_START_BED, None)
        await clear_prompt(context, query.message.chat_id)
        await safe_render_summary(context, query.message.chat_id, date_str)
        return
    if data == "sleep:cancel_wake":
        data = "sleep:toggle"
    if data == "shots:+" or data == "shots:-":
        daily = get_daily_data(context, date_str)
        count = int(parse_sheet_number(daily.get("Стрельнул_раз")))
        if data == "shots:+":
            count += 1
        else:
            count = max(0, count - 1)
        sheets.update_daily_fields(date_str, {COLUMN_MAP["shots"]: count})
        await query.answer()
        await query.edit_message_text(
            f"Стрельнул сегодня: {count}",
            reply_markup=build_shots_keyboard(count),
        )
        return
    if data in {"leisure:sleep", "sleep:toggle"}:
        await query.answer()
        now = get_now(cfg.timezone)
        sleep_start_raw = sheets.get_state(STATE_SLEEP_START)
        if not sleep_start_raw:
            active_day = get_active_date(context)
            sheets.set_state(STATE_SLEEP_START, now.isoformat())
            sheets.set_state(STATE_SLEEP_START_DAY, active_day)
            sheets.set_state(STATE_SLEEP_START_BED, now.strftime("%H:%M"))
            daily = get_daily_data(context, active_day)
            summary = await build_daily_summary(context, active_day)
            await query.edit_message_text(
                f"{summary}\n\n😴 Лег спать. Нажми «Проснулся», когда встанешь.",
                reply_markup=build_main_menu_keyboard(daily),
            )
            return

        try:
            start_dt = datetime.fromisoformat(sleep_start_raw)
        except ValueError:
            start_dt = now
        sleep_day = today_str(cfg.timezone)
        bed_time = sheets.get_state(STATE_SLEEP_START_BED) or start_dt.strftime("%H:%M")
        hours = max(0.0, (now - start_dt).total_seconds() / 3600)
        sheets.update_daily_fields(
            sleep_day,
            {
                COLUMN_MAP["sleep_bed"]: bed_time,
                COLUMN_MAP["sleep_hours"]: f"{hours:.1f}",
                "sleep_source": "manual",
            },
        )
        sheets.set_state(STATE_SLEEP_START, None)
        sheets.set_state(STATE_SLEEP_START_DAY, None)
        sheets.set_state(STATE_SLEEP_START_BED, None)
        sheets.set_state(STATE_ACTIVE_DAY, now.strftime("%Y-%m-%d"))
        new_day = get_active_date(context)
        daily = get_daily_data(context, new_day)
        summary = await build_daily_summary(context, new_day)
        await query.edit_message_text(
            f"{summary}\n\n☀️ Проснулся. Сон: {fmt_num(hours, 1)} ч",
            reply_markup=build_main_menu_keyboard(daily),
        )
        return
    if data == "leisure:sleep":
        daily = get_daily_data(context, date_str)
        current = daily.get("Сон_отбой")
        await show_menu(query, "Сон: во сколько заснул?", mark_set_buttons(SLEEP_BEDTIME_OPTIONS, current), back_to="menu:leisure", cols=3)
        return
    if data == "leisure:productivity":
        daily = get_daily_data(context, date_str)
        current = daily.get("Продуктивность")
        await show_menu(query, "Продуктивность:", mark_set_buttons(PRODUCTIVITY_OPTIONS, current), back_to="menu:leisure", cols=3)
        return
    if data == "leisure:anti":
        await query.answer()
        text, buttons = await build_anti_menu(context, date_str)
        await query.edit_message_text(
            text,
            reply_markup=build_keyboard(buttons, cols=2, back=("⬅️ Назад", "menu:leisure")),
        )
        return

    if data == "food:protein":
        await show_menu(query, "Еда: белковое", FOOD_PROTEIN_OPTIONS, back_to="menu:food", cols=2)
        return
    if data == "food:garnish":
        await show_menu(query, "Еда: гарнир", FOOD_GARNISH_OPTIONS, back_to="menu:food", cols=2)
        return
    if data == "food:sweet":
        await show_menu(query, "Еда: сладкое", FOOD_SWEET_OPTIONS, back_to="menu:food", cols=2)
        return
    if data == "food:oils":
        await show_menu(query, "Еда: масла", FOOD_OIL_OPTIONS, back_to="menu:food", cols=2)
        return
    if data == "food:custom":
        await query.answer()
        context.user_data.clear()
        context.user_data["expect"] = "custom_name"
        await send_or_edit_prompt(
            context,
            query.message.chat_id,
            "Введи название продукта (например, \"Миндаль\").",
            build_keyboard([("⬅️ Назад", "menu:food")], cols=1),
        )
        return

    if data == "morale:mood":
        daily = get_daily_data(context, date_str)
        current = daily.get("Настроение")
        await show_menu(query, "Настроение:", mark_set_buttons(MOOD_OPTIONS, current), back_to="menu:morale", cols=2)
        return
    if data == "morale:energy":
        daily = get_daily_data(context, date_str)
        current = daily.get("Энергия")
        await show_menu(query, "Энергия:", mark_set_buttons(ENERGY_OPTIONS, current), back_to="menu:morale", cols=2)
        return
    if data == "morale:weight":
        await query.answer()
        context.user_data["expect"] = "weight"
        await send_or_edit_prompt(context, query.message.chat_id, "Введи вес (например, 72.4):")
        return
    if data == "morale:regret":
        await query.answer()
        context.user_data["expect"] = "regret"
        await send_or_edit_prompt(context, query.message.chat_id, "О чем жалеешь сегодня? Напиши текст.")
        return
    if data == "morale:review":
        await query.answer()
        context.user_data["expect"] = "review"
        await send_or_edit_prompt(context, query.message.chat_id, "Отзыв о дне: напиши коротко.")
        return

    if data == "habits:text":
        await query.answer()
        context.user_data["expect"] = "habits"
        await send_or_edit_prompt(context, query.message.chat_id, "Привычки: напиши текст.")
        return

    if data.startswith("set:"):
        parts = data.split(":", 2)
        if len(parts) < 3:
            return
        field_key, value = parts[1], parts[2]
        if field_key in FIELD_HEADERS:
            daily = get_daily_data(context, date_str)
            current = daily.get(FIELD_HEADERS[field_key])
            if normalize_choice(current) and normalize_choice(current) != normalize_choice(value):
                next_menu = None
                return_menu = "menu:main"
                if field_key in {"training", "cardio", "steps"}:
                    return_menu = "sport"
                elif field_key in {"english", "ml", "algos", "uni", "code_mode", "code_topic", "reading"}:
                    return_menu = "study"
                elif field_key in {"rest_time", "rest_type", "sleep_bed", "sleep_hours", "sleep_regime", "productivity", "nap"}:
                    return_menu = "leisure"
                elif field_key in {"mood", "energy"}:
                    return_menu = "morale"
                if field_key == "code_mode":
                    next_menu = "code_topic"
                elif field_key == "rest_time":
                    next_menu = "rest_type"
                elif field_key == "sleep_bed":
                    next_menu = "sleep_hours"
                elif field_key == "sleep_hours":
                    next_menu = "sleep_regime"
                await confirm_override(
                    context,
                    query,
                    field_key=field_key,
                    current_value=current,
                    new_value=value,
                    return_menu=return_menu,
                    next_menu=next_menu,
                )
                return
        if field_key == "code_mode":
            sheets.update_daily_fields(date_str, {COLUMN_MAP["code_mode"]: value})
            daily = get_daily_data(context, date_str)
            current_topic = daily.get("Код_тема")
            await show_menu(query, "Код: тема", mark_set_buttons(CODE_TOPIC_OPTIONS, current_topic), back_to="menu:study", cols=2)
            return
        if field_key == "code_topic":
            sheets.update_daily_fields(date_str, {COLUMN_MAP["code_topic"]: value})
            await show_study_menu(query, context, date_str)
            return
        if field_key == "rest_time":
            sheets.update_daily_fields(date_str, {COLUMN_MAP["rest_time"]: value})
            daily = get_daily_data(context, date_str)
            current_type = daily.get("Отдых_тип")
            await show_menu(query, "Отдых: тип", mark_set_buttons(REST_TYPE_OPTIONS, current_type), back_to="menu:leisure", cols=2)
            return
        if field_key == "rest_type":
            sheets.update_daily_fields(date_str, {COLUMN_MAP["rest_type"]: value})
            daily = get_daily_data(context, date_str)
            await show_menu(query, "Досуг:", build_leisure_menu(daily))
            return
        if field_key == "sleep_bed":
            sheets.update_daily_fields(date_str, {COLUMN_MAP["sleep_bed"]: value})
            daily = get_daily_data(context, date_str)
            current_hours = daily.get("Сон_часы")
            await show_menu(query, "Сон: сколько часов?", mark_set_buttons(SLEEP_HOURS_OPTIONS, current_hours), back_to="menu:leisure", cols=3)
            return
        if field_key == "sleep_hours":
            sheets.update_daily_fields(date_str, {COLUMN_MAP["sleep_hours"]: value})
            daily = get_daily_data(context, date_str)
            current_regime = daily.get("Режим")
            await show_menu(query, "Сон: режим", mark_set_buttons(SLEEP_REGIME_OPTIONS, current_regime), back_to="menu:leisure", cols=2)
            return
        if field_key == "sleep_regime":
            sheets.update_daily_fields(date_str, {COLUMN_MAP["sleep_regime"]: value})
            daily = get_daily_data(context, date_str)
            await show_menu(query, "Досуг:", build_leisure_menu(daily))
            return

        field_map = {
            "training": "training",
            "cardio": "cardio",
            "steps": "steps",
            "english": "english",
            "ml": "ml",
            "algos": "algos",
            "uni": "uni",
            "reading": "reading",
            "productivity": "productivity",
            "mood": "mood",
            "energy": "energy",
            "nap": "nap",
        }
        if field_key in field_map:
            key = field_map[field_key]
            col = COLUMN_MAP[key]
            if key == "nap":
                value = float(value)
            elif key in NUMERIC_FIELDS:
                value = int(float(value))
            sheets.update_daily_fields(date_str, {col: value})
            if field_key in {"training", "cardio", "steps"}:
                daily = get_daily_data(context, date_str)
                await show_menu(query, "Спорт:", build_sport_menu(daily))
                return
            if field_key in {"english", "ml", "algos", "uni", "reading"}:
                await show_study_menu(query, context, date_str)
                return
            if field_key in {"productivity", "nap"}:
                daily = get_daily_data(context, date_str)
                await show_menu(query, "Досуг:", build_leisure_menu(daily))
                return
            if field_key in {"mood", "energy"}:
                daily = get_daily_data(context, date_str)
                await show_menu(query, "Моралька:", build_morale_menu(daily))
                return

    if data.startswith("food_item:"):
        portion_code = data.split(":", 1)[1]
        await query.answer()
        await query.edit_message_text("Сколько порций?", reply_markup=quantity_keyboard(portion_code))
        return

    if data.startswith("food_qty:"):
        _, portion_code, qty_str = data.split(":", 2)
        qty = int(qty_str)
        sheets.add_food_log(
            date_str,
            time_str(cfg.timezone),
            portion_code,
            qty)
        await query.answer()
        await query.edit_message_text(
            f"✅ Записал еду: {portion_code} × {qty}",
            reply_markup=build_keyboard(FOOD_MENU, cols=2, back=("⬅️ Назад", "menu:main")),
        )
        return


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_authorized(context, update.effective_user.id if update.effective_user else None):
        return
    expect = context.user_data.get("expect")
    if not expect:
        return
    if update.effective_chat is None or update.message is None:
        return
    cfg = context.application.bot_data["config"]
    sheets = get_sheets(context)
    date_str = get_view_date(context)
    text = update.message.text.strip()
    chat_id = update.effective_chat.id

    if expect == "view_date":
        try:
            picked = datetime.strptime(text, "%Y-%m-%d").date().isoformat()
        except ValueError:
            await send_or_edit_prompt(context, chat_id, "Неверный формат. Используй YYYY-MM-DD")
            return
        set_view_date(context, picked)
        context.user_data.pop("expect", None)
        await finalize_input(context, chat_id, update.message.message_id)
        return

    if expect == "weight":
        try:
            weight = parse_number(text)
        except ValueError:
            await send_or_edit_prompt(context, chat_id, "Не понял вес. Пример: 72.4")
            return
        sheets.update_daily_fields(date_str, {COLUMN_MAP["weight"]: weight})
        context.user_data.clear()
        await finalize_input(context, chat_id, update.message.message_id)
        return

    if expect == "regret":
        sheets.update_daily_fields(date_str, {COLUMN_MAP["regret"]: text})
        context.user_data.clear()
        await finalize_input(context, chat_id, update.message.message_id)
        return

    if expect == "review":
        sheets.update_daily_fields(date_str, {COLUMN_MAP["review"]: text})
        context.user_data.clear()
        await finalize_input(context, chat_id, update.message.message_id)
        return

    if expect == "habits":
        items = parse_habits_value(text)
        sheets.clear_habits_for_date(date_str)
        for item in items:
            if not item:
                continue
            if item not in sheets.get_habits():
                sheets.add_habit(item)
            sheets.set_habit_done(date_str, item, True)
        sheets.update_daily_fields(date_str, {COLUMN_MAP["habits"]: format_habits_value(items)})
        context.user_data.clear()
        await finalize_input(context, chat_id, update.message.message_id)
        return

    if expect == "habit_add":
        added = sheets.add_habit(text)
        context.user_data.clear()
        await finalize_input(context, chat_id, update.message.message_id)
        return

    if expect == "anti_custom":
        reason = text
        sheets.add_session(date_str, time_str(cfg.timezone), "Анти", reason, 0, "")
        context.user_data.clear()
        await finalize_input(context, chat_id, update.message.message_id)
        return

    if expect == "expense_amount":
        category = context.user_data.get("expense_category")
        if not category:
            context.user_data.clear()
            await send_or_edit_prompt(context, chat_id, "Не выбрана категория траты. Открой «Досуг → Траты» ещё раз.")
            return
        try:
            amount = parse_number(text)
        except ValueError:
            await send_or_edit_prompt(context, chat_id, "Нужна сумма числом. Пример: 1450 или 320.5")
            return
        if amount <= 0:
            await send_or_edit_prompt(context, chat_id, "Сумма должна быть больше нуля.")
            return
        sheets.add_expense(date_str, time_str(cfg.timezone), category, amount, "")
        context.user_data.clear()
        await finalize_input(context, chat_id, update.message.message_id)
        return

    if expect == "sleep_bed_manual":
        try:
            hours, minutes = parse_time_hhmm(text)
        except ValueError:
            await send_or_edit_prompt(context, chat_id, "Нужно время в формате HH:MM (например 00:30).")
            return
        bed_time = f"{hours:02d}:{minutes:02d}"
        context.user_data["sleep_bed_manual"] = bed_time
        context.user_data["expect"] = "sleep_hours_manual"
        await send_or_edit_prompt(context, chat_id, "Сколько часов спал? (например 6.5)")
        return

    if expect == "sleep_hours_manual":
        try:
            hours = parse_number(text)
        except ValueError:
            await send_or_edit_prompt(context, chat_id, "Нужны часы числом. Пример: 6.5")
            return
        cfg = context.application.bot_data["config"]
        bed_time = context.user_data.get("sleep_bed_manual")
        if not bed_time:
            context.user_data.clear()
            await send_or_edit_prompt(context, chat_id, "Не нашел время сна, попробуй снова.")
            return
        active_day = get_active_date(context)
        try:
            start_dt = datetime.strptime(f"{active_day} {bed_time}", "%Y-%m-%d %H:%M").replace(
                tzinfo=ZoneInfo(cfg.timezone)
            )
        except Exception:
            start_dt = get_now(cfg.timezone)
        wake_dt = start_dt + timedelta(hours=hours)
        day_shift = (wake_dt.date() - start_dt.date()).days
        suffix = f" (+{day_shift}д)" if day_shift > 0 else ""
        wake_label = wake_dt.strftime("%H:%M") + suffix

        sheets.update_daily_fields(
            active_day,
            {
                COLUMN_MAP["sleep_bed"]: bed_time,
                COLUMN_MAP["sleep_hours"]: f"{hours:.1f}",
                "sleep_source": "manual",
            },
        )
        sheets.set_state(STATE_SLEEP_START, None)
        sheets.set_state(STATE_SLEEP_START_DAY, None)
        sheets.set_state(STATE_SLEEP_START_BED, None)
        sheets.set_state(STATE_ACTIVE_DAY, wake_dt.strftime("%Y-%m-%d"))
        context.user_data.clear()
        await finalize_input(context, chat_id, update.message.message_id)
        return

    if expect == "sleep_bed_edit":
        try:
            hours, minutes = parse_time_hhmm(text)
        except ValueError:
            await send_or_edit_prompt(context, chat_id, "Нужно время в формате HH:MM (например 00:30).")
            return
        cfg = context.application.bot_data["config"]
        now = get_now(cfg.timezone)
        sleep_start_raw = sheets.get_state(STATE_SLEEP_START)
        if not sleep_start_raw:
            context.user_data.clear()
            await send_or_edit_prompt(context, chat_id, "Сон сейчас не запущен. Нажми «Лёг спать».")
            return
        try:
            start_dt = datetime.fromisoformat(sleep_start_raw)
        except ValueError:
            start_dt = now
        candidate = start_dt.replace(hour=hours, minute=minutes, second=0, microsecond=0)
        if candidate < start_dt:
            candidate = candidate + timedelta(days=1)
        if candidate > now:
            await send_or_edit_prompt(context, chat_id, "Время сна не может быть позже текущего.")
            return
        sheets.set_state(STATE_SLEEP_START, candidate.isoformat())
        sheets.set_state(STATE_SLEEP_START_DAY, candidate.strftime("%Y-%m-%d"))
        sheets.set_state(STATE_SLEEP_START_BED, candidate.strftime("%H:%M"))
        context.user_data.clear()
        await finalize_input(context, chat_id, update.message.message_id)
        return

    if expect == "nap_hours":
        try:
            hours = parse_number(text)
        except ValueError:
            await send_or_edit_prompt(context, chat_id, "Нужны часы числом. Пример: 1.5")
            return
        sheets.update_daily_fields(date_str, {COLUMN_MAP["nap"]: hours})
        context.user_data.clear()
        await finalize_input(context, chat_id, update.message.message_id)
        return

    if expect == "custom_name":
        context.user_data["custom_name"] = text
        context.user_data["expect"] = "custom_macros"
        await send_or_edit_prompt(context, chat_id, "Введи Б/Ж/У/Ккал на 100г (4 числа через пробел).")
        return

    if expect == "custom_macros":
        try:
            proteins, fats, carbs, kcal = parse_numbers(text, 4)
        except ValueError:
            await send_or_edit_prompt(context, chat_id, "Нужны 4 числа. Пример: 20 5 10 150")
            return
        context.user_data["custom_macros"] = (proteins, fats, carbs, kcal)
        context.user_data["expect"] = "custom_grams"
        await send_or_edit_prompt(context, chat_id, "Сколько грамм съел? (одно число)")
        return

    if expect == "custom_grams":
        try:
            grams = parse_number(text)
        except ValueError:
            await send_or_edit_prompt(context, chat_id, "Нужны граммы числом. Пример: 120")
            return
        name = context.user_data.get("custom_name", "Продукт")
        proteins, fats, carbs, kcal = context.user_data.get("custom_macros", (0, 0, 0, 0))
        code = f"CUST_{get_now(cfg.timezone).strftime('%y%m%d%H%M%S')}"
        sheets.ensure_food_item(name, proteins, fats, carbs, kcal)
        sheets.ensure_portion(code, name, f"{grams} г (custom)", grams)
        sheets.add_food_log(
            date_str,
            time_str(cfg.timezone),
            code,
            1,
            comment="custom")
        context.user_data.clear()
        await finalize_input(context, chat_id, update.message.message_id)
        return


async def build_daily_summary(context: ContextTypes.DEFAULT_TYPE, date_str: str) -> str:
    cfg = context.application.bot_data["config"]
    calendar_date = today_str(cfg.timezone)
    db = get_sheets(context)
    data = get_daily_data(context, date_str)
    if not data:
        active_day = get_active_date(context)
        date_label = "Сегодня" if date_str == active_day else "Дата"
        return f"📅 {date_label}: {date_str}\nПока нет данных."

    macros = data.get("_macros") or {}
    kcal = macros.get("kcal", 0.0)
    protein = macros.get("protein", 0.0)
    fat = macros.get("fat", 0.0)
    carbs = macros.get("carb", 0.0)

    min_ok, context_min = day_minimum_met(data)
    status = day_completion_status(data)

    active_day = data.get("_active_day")
    date_label = "Сегодня" if str(active_day or "") == date_str else "Дата"
    lines = [f"📅 {date_label}: {date_str}"]

    quality = fmt_value(data.get("Качество_дня"))
    if status == "full":
        quality_prefix = "✅"
    elif status == "partial":
        quality_prefix = "🟧"
    elif status == "empty":
        quality_prefix = "⬜"
    else:
        quality_prefix = "❌"
    lines.append(f"{quality_prefix} Качество дня: {quality}")

    steps_display = fmt_steps(context_min["steps"])
    steps_square = steps_status_square(context_min["steps"])
    lines.append(f"{steps_square} Шаги: {steps_display}")

    sleep_hours = context_min["sleep_hours"]
    sleep_display = "—" if sleep_hours <= 0 else f"{fmt_num(sleep_hours, 1)} ч"
    nap_hours = parse_sheet_number(data.get("Сон_дневной"))
    if nap_hours > 0:
        nap_display = f"{fmt_num(nap_hours, 1)} ч"
        lines.append(f"😴 Сон: {sleep_display} (+дневной {nap_display})")
    else:
        lines.append(f"😴 Сон: {sleep_display}")

    if data.get("Вес") not in (None, ""):
        lines.append(f"⚖️ Вес: {data.get('Вес')}")

    shots_count = parse_sheet_number(data.get("Стрельнул_раз"))
    if shots_count:
        lines.append(f"🎯 Стрельнул: {int(shots_count)}")

    training_value = data.get("Тренировка")
    training_display = display_training(training_value) or (training_value if training_value else "")
    if training_value == "Пропустил":
        training_display = "Пропуск"
    sport_parts = []
    if training_display:
        sport_parts.append(training_display)
    if data.get("Кардио_мин"):
        sport_parts.append(f"кардио {data.get('Кардио_мин')}м")
    sport_line = "—" if not sport_parts else ", ".join(sport_parts)
    lines.append(f"🏋️ Спорт: {sport_line}")

    study_parts = []
    if data.get("Английский_мин"):
        study_parts.append(f"англ {data.get('Английский_мин')}м")
    if data.get("ML_мин"):
        study_parts.append(f"ml {data.get('ML_мин')}м")
    if data.get("Алгосы_мин"):
        study_parts.append(f"алг {data.get('Алгосы_мин')}м")
    if data.get("ВУЗ_мин"):
        study_parts.append(f"вуз {data.get('ВУЗ_мин')}м")
    reading_value = data.get("Чтение_стр")
    if reading_is_set(reading_value):
        label = "не читал" if normalize_choice(reading_value) in {"0", "0.0"} else f"{reading_value} стр"
        study_parts.append(f"чтение {label}")
    study_line = "—" if not study_parts else " · ".join(study_parts)
    lines.append(f"📚 Учеба: {study_line}")

    if any([kcal, protein, fat, carbs]):
        lines.append(
            f"🍽 К {fmt_num(kcal)} | Б {fmt_num(protein)} | Ж {fmt_num(fat)} | У {fmt_num(carbs)}"
        )
    else:
        lines.append("🍽 —")

    expense_total = parse_sheet_number(data.get("Траты_всего"))
    if expense_total > 0:
        expense_parts: list[str] = []
        for label, header in EXPENSE_HEADER_BY_LABEL.items():
            value = parse_sheet_number(data.get(header))
            if value > 0:
                expense_parts.append(f"{label.lower()} {fmt_money(value)}")
        preview = " · ".join(expense_parts[:3])
        if len(expense_parts) > 3:
            preview = f"{preview} +{len(expense_parts) - 3}"
        lines.append(f"💸 Траты: {fmt_money(expense_total)} ₽" + (f" ({preview})" if preview else ""))

    morale_parts = []
    if data.get("Настроение"):
        morale_parts.append(f"настроение {data.get('Настроение')}")
    if data.get("Энергия"):
        morale_parts.append(f"энергия {data.get('Энергия')}")
    if morale_parts:
        lines.append(f"🙂 {', '.join(morale_parts)}")

    anti_sessions = db.get_sessions(date_str, category="Анти")
    if anti_sessions:
        reasons = [s.get("subcategory") for s in anti_sessions if s.get("subcategory")]
        preview = ", ".join(reasons[:3])
        if len(reasons) > 3:
            preview = f"{preview} +{len(reasons) - 3}"
        lines.append(f"🧯 Анти‑прокраст.: {preview}")

    habits_value = data.get("Привычки")
    habits_list = parse_habits_value(habits_value)
    if habits_list:
        try:
            total_habits = len(db.get_habits())
            if total_habits:
                lines.append(f"🧠 Привычки: {len(habits_list)}/{total_habits}")
            else:
                lines.append(f"🧠 Привычки: {', '.join(habits_list)}")
        except Exception:
            lines.append(f"🧠 Привычки: {', '.join(habits_list)}")

    if data.get("О_чем_жалею"):
        lines.append(f"📝 О чем жалею: {data.get('О_чем_жалею')}")
    if data.get("Отзыв_о_дне"):
        lines.append(f"🗒 Отзыв: {data.get('Отзыв_о_дне')}")

    missing = data.get("Не_заполнено")
    if missing not in (None, ""):
        lines.append(f"⚠️ Не хватает для зачета: {missing}")

    if date_str != calendar_date:
        lines.append("🛌 День ещё не закрыт — новый начнётся после «Проснулся».")

    return "\n".join(lines)


async def build_food_summary(context: ContextTypes.DEFAULT_TYPE, date_str: str) -> str:
    db = get_sheets(context)
    data = get_daily_data(context, date_str)
    if not data:
        return "🍽 Еда: сегодня пока нет данных."

    macros = data.get("_macros") or {}
    kcal = macros.get("kcal", 0.0)
    protein = macros.get("protein", 0.0)
    fat = macros.get("fat", 0.0)
    carbs = macros.get("carb", 0.0)

    lines = [
        "🍽 Еда за сегодня",
        f"• Ккал: {fmt_num(kcal)}",
        f"• Б/Ж/У: {fmt_num(protein, 1)} / {fmt_num(fat, 1)} / {fmt_num(carbs, 1)}",
    ]

    # List of foods eaten today
    portions = db.list_portions()
    food_log = db.get_food_log(date_str)
    eaten: dict[str, dict[str, float]] = {}
    eaten_products: set[str] = set()
    for item in food_log:
        label = item["label"]
        qty = item.get("quantity") or 0
        grams = item.get("grams") or 0
        if item.get("product"):
            eaten_products.add(item["product"])
        if label not in eaten:
            eaten[label] = {"qty": 0.0, "grams": 0.0}
        eaten[label]["qty"] += qty
        eaten[label]["grams"] += grams

    if eaten:
        lines.append("")
        lines.append("🧾 Съел сегодня:")
        items = list(eaten.items())
        max_items = 12
        for label, stats in items[:max_items]:
            qty = stats["qty"]
            grams = stats["grams"]
            qty_str = fmt_num(qty, 1) if qty % 1 else str(int(qty))
            line = f"• {label} ×{qty_str}"
            if grams > 0:
                line += f" (~{fmt_num(grams)} г)"
            lines.append(line)
        if len(items) > max_items:
            lines.append(f"… ещё {len(items) - max_items} позиций")

    targets = day_targets(data.get("Тренировка"))
    if not targets:
        lines.append("")
        lines.append("⚠️ Тип дня не задан. Выбери тренировку/отдых в разделе «Спорт»,")
        lines.append("чтобы увидеть цели по КБЖУ.")
        return "\n".join(lines)

    kcal_min, kcal_max = targets["kcal"]
    p_min, p_max = targets["protein"]
    f_min, f_max = targets["fat"]
    c_min, c_max = targets["carb"]

    lines.append("")
    lines.append(f"🎯 Цель ({targets['label']})")
    lines.append(f"• Ккал: {kcal_min}–{kcal_max}")
    lines.append(f"• Б/Ж/У: {p_min}–{p_max} / {f_min}–{f_max} / {c_min}–{c_max}")

    def delta_to_min(value: float, min_val: float) -> float:
        return max(0.0, min_val - value)

    def delta_to_max(value: float, max_val: float) -> float:
        return max(0.0, max_val - value)

    def over_max(value: float, max_val: float) -> float:
        return max(0.0, value - max_val)

    d_kcal_min = delta_to_min(kcal, kcal_min)
    d_p_min = delta_to_min(protein, p_min)
    d_f_min = delta_to_min(fat, f_min)
    d_c_min = delta_to_min(carbs, c_min)

    d_kcal_max = delta_to_max(kcal, kcal_max)
    d_p_max = delta_to_max(protein, p_max)
    d_f_max = delta_to_max(fat, f_max)
    d_c_max = delta_to_max(carbs, c_max)

    over_kcal = over_max(kcal, kcal_max)
    over_p = over_max(protein, p_max)
    over_f = over_max(fat, f_max)
    over_c = over_max(carbs, c_max)

    lines.append("")
    lines.append(
        f"⏳ До минимума: Ккал +{fmt_num(d_kcal_min)} | Б +{fmt_num(d_p_min, 1)} | Ж +{fmt_num(d_f_min, 1)} | У +{fmt_num(d_c_min, 1)}"
    )
    lines.append(
        f"📈 До максимума: Ккал {fmt_num(d_kcal_max)} | Б {fmt_num(d_p_max, 1)} | Ж {fmt_num(d_f_max, 1)} | У {fmt_num(d_c_max, 1)}"
    )
    if any(x > 0 for x in (over_kcal, over_p, over_f, over_c)):
        lines.append(
            f"🚨 Перебор: Ккал +{fmt_num(over_kcal)} | Б +{fmt_num(over_p, 1)} | Ж +{fmt_num(over_f, 1)} | У +{fmt_num(over_c, 1)}"
        )

    target_mid = {
        "kcal": (kcal_min + kcal_max) / 2,
        "protein": (p_min + p_max) / 2,
        "fat": (f_min + f_max) / 2,
        "carb": (c_min + c_max) / 2,
    }
    current_vec = {"kcal": kcal, "protein": protein, "fat": fat, "carb": carbs}
    recommendations = recommend_portions(
        current_vec,
        target_mid,
        portions,
        eaten_products=eaten_products,
        max_items=3,
    )
    if recommendations:
        lines.append("")
        lines.append("🤖 Что лучше добрать сейчас:")
        for item in recommendations:
            m = item["macros"]
            lines.append(
                f"• {item['label']} — +{fmt_num(m['kcal'])} ккал, Б {fmt_num(m['protein'],1)}, Ж {fmt_num(m['fat'],1)}, У {fmt_num(m['carb'],1)}"
            )

    plan = build_plan(current_vec, target_mid, portions, eaten_products=eaten_products, max_steps=4)
    if plan:
        lines.append("")
        lines.append("🍱 Черновик рациона на остаток дня:")
        for item in plan:
            lines.append(f"• {item['label']}")

    return "\n".join(lines)


def _write_sheet(ws, headers: list[str], rows: list[dict]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def build_export_workbook(context: ContextTypes.DEFAULT_TYPE) -> Path:
    cfg = context.application.bot_data["config"]
    db = get_sheets(context)
    export_dir = Path(cfg.export_dir)
    if not export_dir.is_absolute():
        export_dir = BASE_DIR / export_dir
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = get_now(cfg.timezone).strftime("%Y%m%d_%H%M%S")
    xlsx_path = export_dir / f"lifeos_export_{timestamp}.xlsx"

    daily_rows = []
    for date_str in db.get_daily_dates():
        data = get_daily_data(context, date_str)
        row = {header: data.get(header, "") for header in DAILY_HEADERS}
        daily_rows.append(row)

    wb = Workbook()
    ws = wb.active
    ws.title = "daily_summary"
    _write_sheet(ws, DAILY_HEADERS, daily_rows)

    ws = wb.create_sheet("food_log")
    _write_sheet(ws, ["date", "time", "portion_code", "quantity", "comment"], db.list_food_log_all())

    ws = wb.create_sheet("session_log")
    _write_sheet(ws, ["date", "time", "category", "subcategory", "minutes", "comment"], db.list_session_log_all())

    ws = wb.create_sheet("expense_log")
    _write_sheet(ws, ["date", "time", "category", "amount", "comment"], db.list_expense_log_all())

    ws = wb.create_sheet("food_items")
    _write_sheet(ws, ["name", "protein_100", "fat_100", "carb_100", "kcal_100"], db.list_food_items())

    ws = wb.create_sheet("portions")
    _write_sheet(ws, ["code", "product", "description", "grams"], db.list_portions_raw())

    ws = wb.create_sheet("habits")
    _write_sheet(ws, ["id", "name", "active"], db.list_habits_raw())

    ws = wb.create_sheet("habit_log")
    _write_sheet(ws, ["date", "habit", "done"], db.list_habit_log_all())

    wb.save(xlsx_path)
    return xlsx_path


async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_authorized(context, update.effective_user.id if update.effective_user else None):
        return
    if update.message is None:
        return
    xlsx_path = build_export_workbook(context)
    with xlsx_path.open("rb") as f:
        await update.message.reply_document(
            document=f,
            filename=xlsx_path.name,
            caption="Экспорт готов ✅",
        )
    await safe_delete_message(context.bot, update.effective_chat.id, update.message.message_id)


async def static_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_authorized(context, update.effective_user.id if update.effective_user else None):
        return
    if update.message is None:
        return
    await clear_prompt(context, update.effective_chat.id)
    await render_stats(context, update.effective_chat.id, "week")
    await safe_delete_message(context.bot, update.effective_chat.id, update.message.message_id)


def parse_sync_payload(text: str) -> dict:
    parts = text.split(maxsplit=1)
    if len(parts) < 2:
        raise ValueError("Нужен JSON после команды /sync")
    raw = parts[1].strip()
    return json.loads(raw)


def resolve_sync_date(db: Database, cfg, payload_date: object) -> tuple[str, bool]:
    today = today_str(cfg.timezone)
    active_day = db.get_state(STATE_ACTIVE_DAY) or today
    if not payload_date:
        return active_day, True

    incoming = str(payload_date).strip()
    # While active day is not rolled over by wake-up, never let sync jump ahead.
    # Ignore future-day payloads to avoid overwriting active day with "new day zeros".
    if incoming > active_day:
        return active_day, False
    return incoming, True


def apply_sync_payload(db: Database, cfg, payload: dict) -> tuple[str, dict[str, object]]:
    date_str, accepted = resolve_sync_date(db, cfg, payload.get("date"))
    if not accepted:
        LOGGER.info("Ignoring sync payload for future date %s while active_day=%s", payload.get("date"), date_str)
        return date_str, {}
    db.ensure_daily_row(date_str)

    row = db.get_daily_row(date_str) or {}
    updates: dict[str, object] = {}
    if "steps" in payload:
        steps = int(float(payload["steps"]))
        updates[COLUMN_MAP["steps_count"]] = steps
        updates[COLUMN_MAP["steps"]] = steps_to_category(steps)
    if "active_kcal" in payload:
        updates[COLUMN_MAP["active_kcal"]] = float(payload["active_kcal"])
    if "weight" in payload:
        updates[COLUMN_MAP["weight"]] = float(payload["weight"])
    if "sleep_hours" in payload:
        if row.get("sleep_source") != "manual":
            updates[COLUMN_MAP["sleep_hours"]] = str(payload["sleep_hours"])
            updates["sleep_source"] = "health_connect"
    if "english_min" in payload:
        updates[COLUMN_MAP["english"]] = int(float(payload["english_min"]))
    if "ml_min" in payload:
        updates[COLUMN_MAP["ml"]] = int(float(payload["ml_min"]))
    if "algo_min" in payload:
        updates[COLUMN_MAP["algos"]] = int(float(payload["algo_min"]))
    if "algos_min" in payload:
        updates[COLUMN_MAP["algos"]] = int(float(payload["algos_min"]))
    if "uni_min" in payload:
        updates[COLUMN_MAP["uni"]] = int(float(payload["uni_min"]))
    if "nap_hours" in payload:
        updates[COLUMN_MAP["nap"]] = float(payload["nap_hours"])

    food_payload = payload.get("food")
    if isinstance(food_payload, dict):
        if "kcal" in food_payload:
            updates[COLUMN_MAP["food_kcal"]] = float(food_payload.get("kcal", 0))
        if "protein" in food_payload:
            updates[COLUMN_MAP["food_protein"]] = float(food_payload.get("protein", 0))
        if "fat" in food_payload:
            updates[COLUMN_MAP["food_fat"]] = float(food_payload.get("fat", 0))
        if "carb" in food_payload:
            updates[COLUMN_MAP["food_carb"]] = float(food_payload.get("carb", 0))
        updates[COLUMN_MAP["food_tracked"]] = 1
        updates[COLUMN_MAP["food_source"]] = payload.get("food_source", "health_connect")
    elif "food_tracked" in payload:
        updates[COLUMN_MAP["food_tracked"]] = 1 if payload.get("food_tracked") else 0
        if payload.get("food_tracked"):
            updates[COLUMN_MAP["food_source"]] = payload.get("food_source", "health_connect")

    if updates:
        db.update_daily_fields(date_str, updates)
    return date_str, updates


async def sync_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_authorized(context, update.effective_user.id if update.effective_user else None):
        return
    if update.message is None:
        return
    cfg = context.application.bot_data["config"]
    db = get_sheets(context)
    try:
        payload = parse_sync_payload(update.message.text or "")
    except Exception:
        await send_or_edit_prompt(
            context,
            update.effective_chat.id,
            "Не понял /sync. Формат: /sync {\"steps\":12345,...}",
        )
        return
    date_str, updates = apply_sync_payload(db, cfg, payload)
    await safe_delete_message(context.bot, update.effective_chat.id, update.message.message_id)
    await render_summary(context, update.effective_chat.id, date_str)


def start_sync_http_server(db: Database, cfg) -> ThreadingHTTPServer | None:
    token = (cfg.sync_http_token or "").strip()
    if not token:
        LOGGER.info("Sync HTTP disabled (SYNC_HTTP_TOKEN not set).")
        return None

    host = cfg.sync_http_host or "0.0.0.0"
    port = int(cfg.sync_http_port or 8088)

    class SyncHandler(BaseHTTPRequestHandler):
        server_version = "LifeOSSync/1.0"

        def _send_json(self, status: int, payload: dict) -> None:
            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def _get_token(self) -> str | None:
            header_token = self.headers.get("X-Api-Key") or self.headers.get("X-Api-Token")
            if header_token:
                return header_token.strip()
            auth = self.headers.get("Authorization", "")
            if auth.lower().startswith("bearer "):
                return auth[7:].strip()
            return None

        def log_message(self, fmt: str, *args) -> None:
            LOGGER.info("sync-http %s - %s", self.client_address[0], fmt % args)

        def do_GET(self) -> None:
            path = self.path.split("?", 1)[0]
            if path in {"/", "/health"}:
                self._send_json(200, {"ok": True})
                return
            self._send_json(404, {"ok": False, "error": "not_found"})

        def do_POST(self) -> None:
            path = self.path.split("?", 1)[0]
            if path != "/sync":
                self._send_json(404, {"ok": False, "error": "not_found"})
                return

            if self._get_token() != token:
                self._send_json(401, {"ok": False, "error": "unauthorized"})
                return

            try:
                length = int(self.headers.get("Content-Length", "0"))
            except ValueError:
                length = 0
            if length <= 0 or length > 1024 * 1024:
                self._send_json(400, {"ok": False, "error": "invalid_body"})
                return

            raw = self.rfile.read(length)
            try:
                payload = json.loads(raw.decode("utf-8"))
            except Exception:
                self._send_json(400, {"ok": False, "error": "bad_json"})
                return
            if not isinstance(payload, dict):
                self._send_json(400, {"ok": False, "error": "bad_payload"})
                return

            try:
                date_str, updates = apply_sync_payload(db, cfg, payload)
            except Exception:
                LOGGER.exception("Sync HTTP failed")
                self._send_json(500, {"ok": False, "error": "server_error"})
                return

            self._send_json(
                200,
                {
                    "ok": True,
                    "date": date_str,
                    "updated": list(updates.keys()),
                },
            )

    server = ThreadingHTTPServer((host, port), SyncHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    LOGGER.info("Sync HTTP server listening on %s:%s", host, port)
    return server


def main() -> None:
    config = load_config()
    db_path = Path(config.db_path)
    if not db_path.is_absolute():
        db_path = BASE_DIR / db_path
    db = Database(str(db_path))
    db.init_schema()
    db.seed_from_csv(str(BASE_DIR / "data/food_items.csv"), str(BASE_DIR / "data/portions.csv"))

    app = ApplicationBuilder().token(config.telegram_token).build()
    app.bot_data["db"] = db
    app.bot_data["config"] = config
    app.bot_data["allowed_user_id"] = config.allowed_user_id

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("export", export_command))
    app.add_handler(CommandHandler("sync", sync_command))
    app.add_handler(CommandHandler("static", static_command))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_error_handler(handle_error)

    LOGGER.info("Bot started")
    if sys.platform.startswith("win"):
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    sync_server = start_sync_http_server(db, config)
    try:
        app.run_polling()
    finally:
        if sync_server:
            sync_server.shutdown()


if __name__ == "__main__":
    main()
